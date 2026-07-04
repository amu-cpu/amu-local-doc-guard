from __future__ import annotations

import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from uuid import uuid4
from zipfile import ZipFile
from xml.etree import ElementTree as ET

from .excel_preview import read_upload_meta, upload_source_path
from .processor import OUTPUT_ROOT, ProcessingError, sanitize_filename, write_json


EXCEL_ERROR_VALUES = ("#NAME?", "#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NUM!", "#NULL!")
INTERNAL_KEYWORDS = (
    "AI",
    "Codex",
    "内部审稿",
    "内部质检",
    "客户补充回复",
    "用户本轮指定",
    "TODO",
    "待补充",
    "这里可补充",
    "仅供参考",
    "27450",
    "1500万",
    "1500 万",
    "1.5万㎡",
    "1.5 万㎡",
    "天佑城",
    "万民商场",
)


def check_excel_upload_formulas(upload_id: str, output_root: Path = OUTPUT_ROOT) -> dict:
    upload_meta = read_upload_meta(upload_id, output_root)
    source_path = upload_source_path(upload_meta, output_root)
    return check_excel_file(source_path, output_root=output_root, original_name=upload_meta.get("original_name") or source_path.name)


def check_excel_file(path: Path, output_root: Path = OUTPUT_ROOT, original_name: str | None = None) -> dict:
    report = analyze_excel_file(path, original_name=original_name)
    job_id = report["job_id"]
    job_dir = (output_root / job_id).resolve()
    job_dir.mkdir(parents=True, exist_ok=False)
    json_path = job_dir / f"{sanitize_filename(Path(original_name or report['checked_file_name']).stem) or 'workbook'}_公式错误检查报告.json"
    markdown_path = job_dir / f"{sanitize_filename(Path(original_name or report['checked_file_name']).stem) or 'workbook'}_公式错误检查报告.md"
    report["json_report_path"] = str(json_path)
    report["markdown_report_path"] = str(markdown_path)
    report["outputs"] = {"json": json_path.name, "markdown": markdown_path.name}
    write_json(json_path, report)
    markdown_path.write_text(render_markdown_report(report), encoding="utf-8")
    return report


def analyze_excel_file(path: Path, original_name: str | None = None) -> dict:
    started = time.time()
    workbook_path = Path(path).resolve()
    if not workbook_path.exists():
        raise ProcessingError(f"Excel 文件不存在：{workbook_path}")
    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ProcessingError("公式错误检查仅支持 .xlsx、.xlsm；.xls 请先另存为 .xlsx 后再检查。")

    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ProcessingError(f"缺少 openpyxl 依赖，无法检查 Excel：{exc}") from exc

    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    formula_wb = None
    data_wb = None
    safe_stem = sanitize_filename(Path(original_name or workbook_path.name).stem) or "workbook"
    job_id = f"excel_formula_check_{safe_stem}_{uuid4().hex[:8]}"

    try:
        formula_wb = load_workbook(workbook_path, data_only=False, read_only=False, keep_vba=keep_vba)
        data_wb = load_workbook(workbook_path, data_only=True, read_only=False, keep_vba=keep_vba)
        report = build_formula_check_report(formula_wb, data_wb, workbook_path)
    except Exception as exc:
        raise ProcessingError(f"读取或检查 Excel 失败：{exc}") from exc
    finally:
        if formula_wb is not None:
            formula_wb.close()
        if data_wb is not None:
            data_wb.close()

    report.update(
        {
            "job_id": job_id,
            "checked_file_path": str(workbook_path),
            "checked_file_name": workbook_path.name,
            "check_time": datetime.now().astimezone().isoformat(timespec="seconds"),
            "duration_seconds": round(time.time() - started, 2),
            "recalculation": {
                "attempted": False,
                "supported": None,
                "message": "未执行本机重算，仅完成静态扫描和缓存值扫描。",
            },
        }
    )
    report["passed"] = not report["has_formula_errors"]
    report["final_conclusion"] = "通过" if report["passed"] else "不通过"
    report["has_internal_keyword_hits"] = bool(report.get("internal_keyword_hits"))
    report["minimum_fix_targets"] = build_minimum_fix_targets(report)
    return report


def build_formula_check_report(formula_wb, data_wb, workbook_path: Path) -> dict:
    error_records: list[dict] = []
    keyword_hits: list[dict] = []
    sheet_summaries = []
    formula_cells_count = 0
    sheet_states = Counter()

    for formula_sheet in formula_wb.worksheets:
        state = getattr(formula_sheet, "sheet_state", "visible") or "visible"
        sheet_states[state] += 1
        data_sheet = data_wb[formula_sheet.title] if formula_sheet.title in data_wb.sheetnames else None
        sheet_error_counter: Counter[str] = Counter()
        sheet_formula_count = 0

        for row in formula_sheet.iter_rows():
            for cell in row:
                address = cell.coordinate
                formula_value = cell.value
                cached_value = data_sheet[address].value if data_sheet is not None else None
                if is_formula(formula_value):
                    sheet_formula_count += 1
                    formula_cells_count += 1
                    scan_values = [("formula", formula_value), ("cached_value", cached_value)]
                else:
                    scan_values = [("cell_value", formula_value)]

                for value_source, value in scan_values:
                    for error_value in find_error_tokens(value):
                        error_records.append(
                            {
                                "sheet_name": formula_sheet.title,
                                "sheet_state": state,
                                "cell": address,
                                "error_value": error_value,
                                "value_source": value_source,
                                "formula": formula_value if is_formula(formula_value) else "",
                                "display_or_cached_value": stringify(cached_value if is_formula(formula_value) else formula_value),
                                "row_hidden": bool(formula_sheet.row_dimensions[cell.row].hidden),
                                "column_hidden": bool(formula_sheet.column_dimensions[cell.column_letter].hidden),
                            }
                        )
                        sheet_error_counter[error_value] += 1

                scan_text_for_keywords(keyword_hits, formula_sheet.title, address, "cell", formula_value)
                if is_formula(formula_value):
                    scan_text_for_keywords(keyword_hits, formula_sheet.title, address, "cached_value", cached_value)
                if cell.comment:
                    scan_text_for_keywords(keyword_hits, formula_sheet.title, address, "comment", cell.comment.text)
                if cell.hyperlink:
                    hyperlink_value = getattr(cell.hyperlink, "target", "") or getattr(cell.hyperlink, "display", "")
                    scan_text_for_keywords(keyword_hits, formula_sheet.title, address, "hyperlink", hyperlink_value)

        scan_text_for_keywords(keyword_hits, formula_sheet.title, "", "sheet_name", formula_sheet.title)
        sheet_summaries.append(
            {
                "sheet_name": formula_sheet.title,
                "state": state,
                "formula_cells_count": sheet_formula_count,
                "error_counts": dict(sheet_error_counter),
                "total_errors": sum(sheet_error_counter.values()),
            }
        )

    defined_name_issues, defined_name_keyword_hits = inspect_defined_names(formula_wb)
    keyword_hits.extend(defined_name_keyword_hits)
    external_links = inspect_external_links(workbook_path, formula_wb)

    error_counts = Counter(record["error_value"] for record in error_records)
    name_error_by_sheet = defaultdict(int)
    other_formula_errors_by_sheet = defaultdict(int)
    for record in error_records:
        if record["error_value"] == "#NAME?":
            name_error_by_sheet[record["sheet_name"]] += 1
        else:
            other_formula_errors_by_sheet[record["sheet_name"]] += 1

    return {
        "worksheet_total_count": len(formula_wb.worksheets),
        "visible_sheet_count": sheet_states.get("visible", 0),
        "hidden_sheet_count": sheet_states.get("hidden", 0) + sheet_states.get("veryHidden", 0),
        "very_hidden_sheet_count": sheet_states.get("veryHidden", 0),
        "sheet_summaries": sheet_summaries,
        "formula_cells_count": formula_cells_count,
        "has_formula_errors": bool(error_records),
        "error_counts": {error: error_counts.get(error, 0) for error in EXCEL_ERROR_VALUES},
        "total_error_count": len(error_records),
        "name_error_count_by_sheet": dict(name_error_by_sheet),
        "other_formula_error_count_by_sheet": dict(other_formula_errors_by_sheet),
        "error_locations": error_records,
        "internal_keyword_hits": keyword_hits,
        "has_blocking_internal_keywords": False,
        "external_links": external_links,
        "defined_name_issues": defined_name_issues,
    }


def is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def stringify(value) -> str:
    if value is None:
        return ""
    return str(value)


def find_error_tokens(value) -> list[str]:
    if value is None:
        return []
    text = str(value)
    return [error for error in EXCEL_ERROR_VALUES if error in text]


def scan_text_for_keywords(hits: list[dict], sheet_name: str, cell: str, source: str, value) -> None:
    if value is None:
        return
    text = str(value)
    for keyword in INTERNAL_KEYWORDS:
        if keyword in text:
            hits.append({"sheet_name": sheet_name, "cell": cell, "source": source, "keyword": keyword})


def inspect_defined_names(workbook) -> tuple[list[dict], list[dict]]:
    issues: list[dict] = []
    keyword_hits: list[dict] = []
    defined_names = getattr(workbook, "defined_names", None)
    if not defined_names:
        return issues, keyword_hits

    try:
        iterable = defined_names.items()
    except Exception:
        iterable = []

    for name, defined_name in iterable:
        attr_text = stringify(getattr(defined_name, "attr_text", ""))
        if not attr_text:
            issues.append({"name": name, "issue": "empty_reference", "error_value": "", "reference": attr_text})
        lower_text = attr_text.lower()
        if "[" in attr_text or "http://" in lower_text or "https://" in lower_text:
            issues.append({"name": name, "issue": "external_reference", "error_value": "", "reference": attr_text})
        if attr_text.startswith("{") and attr_text.endswith("}"):
            continue
        if "#REF!" in attr_text:
            issues.append({"name": name, "issue": "error_token", "error_value": "#REF!", "reference": attr_text})
        try:
            list(defined_name.destinations)
        except Exception as exc:
            if attr_text and attr_text not in {"", "#N/A"}:
                issues.append({"name": name, "issue": "unresolved_reference", "error_value": "", "reference": attr_text, "message": str(exc)})
        scan_text_for_keywords(keyword_hits, "", "", f"defined_name:{name}", attr_text)
    return issues, keyword_hits


def inspect_external_links(workbook_path: Path, workbook) -> list[dict]:
    links: list[dict] = []
    seen = set()

    for link in getattr(workbook, "_external_links", []) or []:
        target = stringify(getattr(link, "file_link", "") or getattr(link, "target", "") or link)
        if target and target not in seen:
            links.append({"source": "openpyxl", "target": target})
            seen.add(target)

    try:
        with ZipFile(workbook_path) as archive:
            for name in archive.namelist():
                if name.startswith("xl/externalLinks/") or "externalLink" in name:
                    if name not in seen:
                        links.append({"source": "xlsx_package", "target": name})
                        seen.add(name)
                if name.endswith(".rels"):
                    content = archive.read(name).decode("utf-8", errors="ignore")
                    for match in re_find_target(content):
                        if match not in seen:
                            links.append({"source": name, "target": match})
                            seen.add(match)
    except Exception as exc:
        links.append({"source": "xlsx_package", "target": "", "error": str(exc)})
    return links


def re_find_target(content: str) -> list[str]:
    results = []
    start = 0
    token = 'Target="'
    while True:
        idx = content.find(token, start)
        if idx < 0:
            break
        begin = idx + len(token)
        end = content.find('"', begin)
        if end < 0:
            break
        target = content[begin:end]
        if ("externalLink" in target) or _looks_like_external_target(target):
            results.append(target)
        start = end + 1
    return results


def _looks_like_external_target(target: str) -> bool:
    lower = target.lower()
    return lower.startswith("http://") or lower.startswith("https://") or lower.startswith("file:") or (len(target) > 2 and target[1:3] == ":\\") or target.startswith("\\\\")


def build_minimum_fix_targets(report: dict) -> list[dict]:
    targets = {}
    for record in report.get("error_locations", []):
        sheet_name = record["sheet_name"]
        targets.setdefault(sheet_name, [])
        if len(targets[sheet_name]) < 20:
            targets[sheet_name].append(record["cell"])
    return [{"sheet_name": sheet, "cells": cells} for sheet, cells in targets.items()]


def render_markdown_report(report: dict) -> str:
    lines = [
        "# Excel 交付前质检 / 公式错误扫描报告",
        "",
        f"- 检查文件路径：`{report['checked_file_path']}`",
        f"- 检查时间：{report['check_time']}",
        f"- 工作表总数：{report['worksheet_total_count']}",
        f"- 可见表数量：{report['visible_sheet_count']}",
        f"- 隐藏表数量：{report['hidden_sheet_count']}（veryHidden：{report['very_hidden_sheet_count']}）",
        f"- 公式单元格数量：{report['formula_cells_count']}",
        f"- 是否存在公式错误：{'是' if report['has_formula_errors'] else '否'}",
        f"- 最终结论：{report['final_conclusion']}",
        f"- 重算状态：{report['recalculation']['message']}",
        "",
        "## 各类错误数量",
        "",
    ]
    for error in EXCEL_ERROR_VALUES:
        lines.append(f"- `{error}`：{report['error_counts'].get(error, 0)}")

    lines.extend(["", "## 错误位置清单", ""])
    if report.get("error_locations"):
        lines.append("| 工作表 | 单元格 | 错误值 | 来源 | 公式或显示值 |")
        lines.append("|---|---:|---:|---|---|")
        for record in report["error_locations"]:
            value = record.get("formula") or record.get("display_or_cached_value") or ""
            lines.append(
                f"| {escape_md(record['sheet_name'])} | `{record['cell']}` | `{record['error_value']}` | "
                f"{escape_md(record['value_source'])} | `{escape_md(value)}` |"
            )
    else:
        lines.append("未发现指定 Excel 错误值。")

    lines.extend(["", "## 内部痕迹关键词命中清单", ""])
    if report.get("internal_keyword_hits"):
        lines.append("| 工作表 | 单元格 | 来源 | 关键词 |")
        lines.append("|---|---:|---|---|")
        for hit in report["internal_keyword_hits"]:
            lines.append(f"| {escape_md(hit.get('sheet_name', ''))} | `{hit.get('cell', '')}` | {escape_md(hit.get('source', ''))} | `{escape_md(hit['keyword'])}` |")
    else:
        lines.append("未发现指定内部痕迹关键词。")

    lines.extend(["", "## 外部链接清单", ""])
    if report.get("external_links"):
        for link in report["external_links"]:
            lines.append(f"- {escape_md(link.get('source', ''))}：`{escape_md(link.get('target', ''))}`")
    else:
        lines.append("未发现外部链接。")

    lines.extend(["", "## 命名区域异常清单", ""])
    if report.get("defined_name_issues"):
        for issue in report["defined_name_issues"]:
            lines.append(f"- {escape_md(issue.get('name', ''))}：{escape_md(issue.get('issue', ''))} `{escape_md(issue.get('reference', ''))}`")
    else:
        lines.append("未发现命名区域异常。")

    lines.extend(["", "## 最少需要修复的位置", ""])
    if report.get("minimum_fix_targets"):
        for target in report["minimum_fix_targets"]:
            cells = ", ".join(f"`{cell}`" for cell in target["cells"])
            lines.append(f"- {escape_md(target['sheet_name'])}：{cells}")
    else:
        lines.append("无。")
    lines.append("")
    return "\n".join(lines)


def escape_md(value) -> str:
    return stringify(value).replace("|", "\\|").replace("\n", " ")
