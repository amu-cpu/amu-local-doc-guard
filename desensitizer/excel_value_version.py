from __future__ import annotations

import shutil
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from uuid import uuid4

from .excel_formula_check import EXCEL_ERROR_VALUES, analyze_excel_file, render_markdown_report
from .excel_preview import (
    XL_PASTE_VALUES,
    XL_SHEET_VISIBLE,
    calculate_workbook,
    cleanup_com_application,
    cleanup_new_office_processes,
    com_app_process_ids,
    count_formulas,
    dedupe_texts,
    find_windows_desktop,
    looks_like_locked_file_error,
    office_process_snapshot,
    open_com_workbook,
    read_upload_meta,
    report_progress,
    upload_source_path,
)
from .processor import OUTPUT_ROOT, ProcessingError, sanitize_filename, write_json


INTERNAL_SHEET_KEYWORDS = (
    "模型检查",
    "缓存检查",
    "临时表",
    "调试表",
    "测试表",
    "内部",
    "过程",
)
DEFINED_NAME_BAD_KEYWORDS = (
    "#REF!",
    "旧模板",
    "乱码",
    "old_template",
    "oldtemplate",
    "template_old",
)
MODEL_VERSION_MARKER = "模型版"
VALUE_VERSION_MARKER = "客户交付值版"


def generate_value_version_from_upload(
    upload_id: str,
    output_root: Path = OUTPUT_ROOT,
    keep_hidden_sheets: bool = False,
    protection_password: str = "123456",
    publish_to_desktop: bool = True,
    progress_callback=None,
) -> dict:
    upload_meta = read_upload_meta(upload_id, output_root)
    source_path = upload_source_path(upload_meta, output_root)
    original_name = sanitize_filename(upload_meta.get("original_name") or source_path.name)
    return generate_value_version_from_path(
        source_path,
        output_root=output_root,
        original_name=original_name,
        keep_hidden_sheets=keep_hidden_sheets,
        protection_password=protection_password,
        publish_to_desktop=publish_to_desktop,
        progress_callback=progress_callback,
    )


def generate_value_version_from_path(
    source_path: Path,
    output_root: Path = OUTPUT_ROOT,
    original_name: str | None = None,
    output_dir: Path | None = None,
    keep_hidden_sheets: bool = False,
    protection_password: str = "123456",
    publish_to_desktop: bool = False,
    allow_value_as_source: bool = False,
    progress_callback=None,
) -> dict:
    started = time.time()
    source_input_path = Path(source_path).expanduser()
    source_absolute_path = source_input_path.absolute()
    source_display_path = str(source_absolute_path)
    source_path = source_input_path.resolve()
    if not source_path.exists():
        raise ProcessingError(f"Excel 文件不存在：{source_display_path}")
    original_name = sanitize_filename(original_name or source_path.name)
    suffix = source_path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise ProcessingError("客户交付值版生成仅支持 .xlsx、.xlsm；.xls 请先另存为 .xlsx 后再处理。")

    source_info = build_source_version_info(original_name)
    if source_info["is_value_version"] and not allow_value_as_source:
        raise ProcessingError("当前输入疑似已是客户交付值版，请提供模型版或原始带公式版。")

    safe_stem = source_info["base_stem"]
    job_id = f"excel_value_version_{safe_stem}_{uuid4().hex[:8]}"
    if output_dir is None:
        excel_dir = (Path(output_root).expanduser() / job_id).absolute()
        report_dir = excel_dir
        excel_dir.mkdir(parents=True, exist_ok=False)
        output_mode = "desktop" if publish_to_desktop else "app_output"
    else:
        excel_dir = Path(output_dir).expanduser()
        excel_dir.mkdir(parents=True, exist_ok=True)
        report_dir = (Path(output_root).expanduser() / job_id).absolute()
        report_dir.mkdir(parents=True, exist_ok=False)
        output_mode = "source_dir"

    model_suffix = ".xlsm" if suffix == ".xlsm" else ".xlsx"
    should_copy_model = not source_info["is_model_version"] or output_dir is None
    model_path = (
        unique_child_path(excel_dir, f"{safe_stem}_模型版{model_suffix}")
        if should_copy_model
        else source_absolute_path
    )
    value_path = unique_child_path(excel_dir, f"{safe_stem}_客户交付值版.xlsx")
    json_report_path = unique_child_path(report_dir, f"{safe_stem}_客户交付值版_质检报告.json")
    markdown_report_path = unique_child_path(report_dir, f"{safe_stem}_客户交付值版_质检报告.md")

    warnings: list[str] = []
    errors: list[str] = []
    if should_copy_model:
        report_progress(progress_callback, "正在复制模型版...", 0, None)
        shutil.copy2(source_path, model_path)
    else:
        report_progress(progress_callback, "正在读取模型版...", 0, None)
        warnings.append("输入文件已是模型版，同目录快链路未重复生成模型版。")

    source_check_started = time.time()
    source_model_check = analyze_excel_file(model_path, original_name=model_path.name)
    source_check_seconds = round(time.time() - source_check_started, 2)
    if suffix == ".xlsm":
        warnings.append("源文件是 .xlsm，模型版保留 .xlsm 后缀；客户交付值版导出为 .xlsx。")

    report_progress(progress_callback, "正在生成客户交付值版...", 1, None)
    generation_summary = None
    conversion_errors = []
    working_path = model_path
    if suffix == ".xlsx":
        shutil.copy2(model_path, value_path)
        working_path = value_path

    conversion_started = time.time()
    for prefer_wps in (False, True):
        try:
            generation_summary = convert_with_com(
                working_path,
                value_path,
                keep_hidden_sheets=keep_hidden_sheets,
                protection_password=protection_password,
                prefer_wps=prefer_wps,
                progress_callback=progress_callback,
            )
            break
        except ProcessingError as exc:
            conversion_errors.append(str(exc))

    if generation_summary is None:
        try:
            warnings.extend(conversion_errors)
            generation_summary = convert_with_openpyxl(
                model_path,
                value_path,
                keep_hidden_sheets=keep_hidden_sheets,
                progress_callback=progress_callback,
            )
        except ProcessingError as exc:
            raise ProcessingError("客户交付值版生成失败：" + "；".join(conversion_errors + [str(exc)])) from exc
    conversion_seconds = round(time.time() - conversion_started, 2)

    report_progress(progress_callback, "正在复检客户交付值版...", 2, None)
    delivery_check_started = time.time()
    delivery_value_check = analyze_excel_file(value_path, original_name=value_path.name)
    delivery_recheck_seconds = round(time.time() - delivery_check_started, 2)
    strict_result = build_delivery_strict_result(delivery_value_check)
    total_seconds = round(time.time() - started, 2)
    timing_summary = build_timing_summary(
        source_check_seconds=source_check_seconds,
        conversion_seconds=conversion_seconds,
        delivery_recheck_seconds=delivery_recheck_seconds,
        generation_summary=generation_summary,
        total_seconds=total_seconds,
    )

    report = {
        "job_id": job_id,
        "source_file": original_name,
        "source_file_path": source_display_path,
        "source_file_role": source_info["role"],
        "source_name_contains_model_version": source_info["is_model_version"],
        "source_name_contains_value_version": source_info["is_value_version"],
        "is_value_version_used_as_source": source_info["is_value_version"],
        "allow_value_as_source": bool(allow_value_as_source),
        "value_source_misuse_judgement": (
            "allowed_by_explicit_flag"
            if source_info["is_value_version"] and allow_value_as_source
            else "blocked_by_default"
            if source_info["is_value_version"]
            else "not_value_version_source"
        ),
        "output_mode": output_mode,
        "app_output_dir": str(Path(report_dir).absolute()),
        "excel_output_dir": str(Path(excel_dir).absolute()),
        "report_output_dir": str(Path(report_dir).absolute()),
        "model_version_path": str(Path(model_path).absolute()),
        "delivery_value_path": str(Path(value_path).absolute()),
        "json_report_path": str(Path(json_report_path).absolute()),
        "markdown_report_path": str(Path(markdown_report_path).absolute()),
        "model_version_published_path": str(Path(model_path).absolute()),
        "delivery_value_published_path": str(Path(value_path).absolute()),
        "json_report_published_path": str(Path(json_report_path).absolute()),
        "markdown_report_published_path": str(Path(markdown_report_path).absolute()),
        "actual_save_location": str(Path(value_path).absolute()),
        "model_version_file": model_path.name,
        "delivery_value_file": value_path.name,
        "json_report_file": json_report_path.name,
        "markdown_report_file": markdown_report_path.name,
        "outputs": {
            "model": model_path.name,
            "value": value_path.name,
            "json": json_report_path.name,
            "markdown": markdown_report_path.name,
        },
        "keep_hidden_sheets": bool(keep_hidden_sheets),
        "source_model_check": source_model_check,
        "delivery_value_check": delivery_value_check,
        "generation_summary": generation_summary,
        "delivery_ready": strict_result["passed"],
        "blocking_issues": strict_result["blocking_issues"],
        "final_conclusion": "通过，可用于交付" if strict_result["passed"] else "不通过，不建议交付",
        "warnings": dedupe_texts(warnings + generation_summary.get("warnings", [])),
        "errors": errors + generation_summary.get("errors", []),
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "duration_seconds": total_seconds,
        "timing_summary": timing_summary,
        "source_check_seconds": source_check_seconds,
        "conversion_seconds": conversion_seconds,
        "delivery_recheck_seconds": delivery_recheck_seconds,
        "local_only": True,
    }
    desktop_plan = prepare_desktop_publish(
        report,
        model_path=model_path,
        value_path=value_path,
        json_report_path=json_report_path,
        markdown_report_path=markdown_report_path,
        enabled=publish_to_desktop and output_dir is None,
    )
    write_value_version_report_files(report, json_report_path, markdown_report_path)
    publish_value_version_files(report, desktop_plan)
    if report.get("desktop_output_enabled") and report.get("json_report_desktop_path") and report.get("markdown_report_desktop_path"):
        report["desktop_json_copied"] = True
        report["desktop_markdown_copied"] = True
    write_value_version_report_files(report, json_report_path, markdown_report_path)
    if report.get("desktop_output_enabled") and report.get("json_report_desktop_path") and report.get("markdown_report_desktop_path"):
        write_value_version_report_files(
            report,
            Path(report["json_report_desktop_path"]),
            Path(report["markdown_report_desktop_path"]),
        )
    return report


def build_source_version_info(original_name: str) -> dict:
    raw_stem = Path(original_name).stem
    is_value_version = VALUE_VERSION_MARKER in raw_stem
    is_model_version = MODEL_VERSION_MARKER in raw_stem and not is_value_version
    base_stem = raw_stem
    for marker in (VALUE_VERSION_MARKER, MODEL_VERSION_MARKER):
        base_stem = base_stem.replace(marker, "")
    base_stem = cleanup_version_stem(base_stem)
    safe_base_stem = sanitize_filename(base_stem) or "workbook"
    if is_value_version:
        role = "value_version"
    elif is_model_version:
        role = "model_version"
    else:
        role = "raw_or_formula_workbook"
    return {
        "raw_stem": raw_stem,
        "base_stem": safe_base_stem,
        "is_model_version": is_model_version,
        "is_value_version": is_value_version,
        "role": role,
    }


def cleanup_version_stem(stem: str) -> str:
    cleaned = stem
    for token in ("__", "--", "  "):
        while token in cleaned:
            cleaned = cleaned.replace(token, token[0])
    return cleaned.strip(" _-　")


def build_timing_summary(
    source_check_seconds: float,
    conversion_seconds: float,
    delivery_recheck_seconds: float,
    generation_summary: dict,
    total_seconds: float,
) -> dict:
    engine = generation_summary.get("processing_engine", "")
    is_com = "COM" in engine
    recalculation_seconds = generation_summary.get("recalculation_seconds")
    value_conversion_seconds = generation_summary.get("value_conversion_seconds")
    return {
        "source_file_check_seconds": source_check_seconds,
        "excel_com_recalculation_seconds": recalculation_seconds if is_com else 0,
        "value_conversion_seconds": value_conversion_seconds if value_conversion_seconds is not None else conversion_seconds,
        "delivery_recheck_seconds": delivery_recheck_seconds,
        "total_seconds": total_seconds,
    }


def unique_child_path(directory: Path, filename: str) -> Path:
    directory = Path(directory)
    candidate = directory / filename
    if not candidate.exists():
        return candidate
    stem = candidate.stem
    suffix = candidate.suffix
    for index in range(2, 1000):
        versioned = directory / f"{stem}_v{index}{suffix}"
        if not versioned.exists():
            return versioned
    raise ProcessingError(f"无法生成唯一文件名，请清理旧输出后重试：{directory}")


def prepare_desktop_publish(
    report: dict,
    model_path: Path,
    value_path: Path,
    json_report_path: Path,
    markdown_report_path: Path,
    enabled: bool,
) -> list[tuple[Path, Path, str]]:
    report.update(
        {
            "desktop_output_enabled": bool(enabled),
            "desktop_output_dir": "",
            "desktop_model_copied": False,
            "desktop_value_copied": False,
            "desktop_json_copied": False,
            "desktop_markdown_copied": False,
            "desktop_output_message": "",
        }
    )
    if not enabled:
        return []

    desktop_dir = find_windows_desktop()
    if not desktop_dir:
        report["output_mode"] = "app_output"
        report.setdefault("warnings", []).append("桌面目录未找到，已保存在程序 output 目录。")
        return []

    desktop_dir.mkdir(parents=True, exist_ok=True)
    desktop_model = unique_child_path(desktop_dir, model_path.name)
    desktop_value = unique_child_path(desktop_dir, value_path.name)
    desktop_json = unique_child_path(desktop_dir, json_report_path.name)
    desktop_markdown = unique_child_path(desktop_dir, markdown_report_path.name)

    report.update(
        {
            "desktop_output_dir": str(desktop_dir),
            "model_desktop_path": str(desktop_model),
            "delivery_value_desktop_path": str(desktop_value),
            "json_report_desktop_path": str(desktop_json),
            "markdown_report_desktop_path": str(desktop_markdown),
            "model_version_published_path": str(desktop_model),
            "delivery_value_published_path": str(desktop_value),
            "json_report_published_path": str(desktop_json),
            "markdown_report_published_path": str(desktop_markdown),
            "actual_save_location": str(desktop_value),
            "desktop_output_message": f"已默认复制到桌面：{desktop_value.name}",
        }
    )
    return [
        (model_path, desktop_model, "desktop_model_copied"),
        (value_path, desktop_value, "desktop_value_copied"),
    ]


def publish_value_version_files(
    report: dict,
    copy_plan: list[tuple[Path, Path, str]],
) -> None:
    if not copy_plan:
        return

    for source, target, flag in copy_plan:
        try:
            shutil.copy2(source, target)
            report[flag] = True
        except Exception as exc:
            report[flag] = False
            report.setdefault("warnings", []).append(f"复制到桌面失败：{target}。原因：{exc}")


def write_value_version_report_files(report: dict, json_report_path: Path, markdown_report_path: Path) -> None:
    write_json(json_report_path, report)
    markdown_report_path.write_text(render_value_version_report(report), encoding="utf-8")


def convert_with_com(
    workbook_path: Path,
    final_path: Path,
    keep_hidden_sheets: bool,
    protection_password: str,
    prefer_wps: bool,
    progress_callback=None,
) -> dict:
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        raise ProcessingError(f"缺少 Excel/WPS 自动化依赖 pywin32：{exc}") from exc

    engine_name = "WPS COM" if prefer_wps else "Excel COM"
    app = None
    workbook = None
    app_process_ids: set[int] = set()
    office_processes_before = office_process_snapshot()
    pythoncom.CoInitialize()
    try:
        app = win32com.client.DispatchEx("Ket.Application" if prefer_wps else "Excel.Application")
        app_process_ids = com_app_process_ids(app)
        app.Visible = False
        app.DisplayAlerts = False
        try:
            app.ScreenUpdating = False
            app.EnableEvents = False
            app.AskToUpdateLinks = False
        except Exception:
            pass

        workbook = open_com_workbook(app, workbook_path, read_only=False)
        try_unprotect_workbook(workbook, protection_password)
        recalculation_started = time.time()
        calculate_workbook(app)
        try:
            app.CalculateUntilAsyncQueriesDone()
        except Exception:
            pass
        recalculation_seconds = round(time.time() - recalculation_started, 2)

        value_conversion_started = time.time()
        summary = clean_com_workbook(
            workbook,
            app,
            keep_hidden_sheets=keep_hidden_sheets,
            protection_password=protection_password,
            progress_callback=progress_callback,
        )
        value_conversion_seconds = round(time.time() - value_conversion_started, 2)
        if Path(workbook_path).resolve() == Path(final_path).resolve():
            workbook.Save()
        else:
            workbook.SaveAs(str(Path(final_path).resolve()), FileFormat=51)
        summary.update(
            {
                "processing_engine": engine_name,
                "formula_conversion_engine": engine_name,
                "recalculation": {
                    "attempted": True,
                    "supported": True,
                    "message": f"已通过 {engine_name} 执行本机全量重算、保存后再扫描。",
                },
                "recalculation_seconds": recalculation_seconds,
                "value_conversion_seconds": value_conversion_seconds,
            }
        )
        return summary
    except Exception as exc:
        message = str(exc)
        if looks_like_locked_file_error(message):
            raise ProcessingError("请先关闭正在打开的 Excel 文件后再处理。") from exc
        raise ProcessingError(f"{engine_name} 生成客户交付值版失败：{message}") from exc
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if app is not None:
            cleanup_com_application(app, app_process_ids)
        cleanup_new_office_processes(office_processes_before)
        pythoncom.CoUninitialize()


def clean_com_workbook(
    workbook,
    app,
    keep_hidden_sheets: bool,
    protection_password: str,
    progress_callback=None,
) -> dict:
    warnings: list[str] = []
    errors: list[str] = []
    sheet_infos = read_com_sheet_infos(workbook)
    delete_sheet_names = choose_sheets_to_delete(sheet_infos, keep_hidden_sheets)
    deleted_sheets = []
    skipped_delete_sheets = []

    for sheet_name in reversed(delete_sheet_names):
        try:
            sheet = workbook.Worksheets(sheet_name)
            try:
                sheet.Visible = XL_SHEET_VISIBLE
            except Exception:
                pass
            sheet.Delete()
            deleted_sheets.append(sheet_name)
        except Exception as exc:
            skipped_delete_sheets.append({"sheet_name": sheet_name, "message": str(exc)})

    formulas_detected_count = 0
    formulas_converted_count = 0
    processed_sheets = []
    total_sheets = int(workbook.Worksheets.Count)
    for index in range(1, total_sheets + 1):
        sheet = workbook.Worksheets(index)
        sheet_name = str(sheet.Name)
        processed_sheets.append(sheet_name)
        report_progress(progress_callback, f"正在转值工作表：{sheet_name}", index, total_sheets)
        try:
            try_unprotect_sheet(sheet, protection_password)
            used_range = sheet.UsedRange
            formulas_count = count_formulas(used_range)
            formulas_detected_count += formulas_count
            if formulas_count:
                used_range.Copy()
                used_range.PasteSpecial(Paste=XL_PASTE_VALUES)
                app.CutCopyMode = False
                formulas_converted_count += formulas_count
        except Exception as exc:
            errors.append(f"{sheet_name}: {exc}")

    broken_external_links = break_com_external_links(workbook)
    deleted_defined_names = delete_com_defined_names(workbook)
    total_sheets_after_cleanup = int(workbook.Worksheets.Count)
    if skipped_delete_sheets:
        warnings.append("部分内部或隐藏工作表未能自动删除，请查看报告后人工处理。")

    return {
        "total_sheets_after_cleanup": total_sheets_after_cleanup,
        "processed_sheets": processed_sheets,
        "processed_sheets_count": len(processed_sheets),
        "deleted_sheets": deleted_sheets,
        "deleted_sheets_count": len(deleted_sheets),
        "skipped_delete_sheets": skipped_delete_sheets,
        "formulas_detected_count": formulas_detected_count,
        "formulas_converted_count": formulas_converted_count,
        "formula_removed": formulas_detected_count == formulas_converted_count,
        "broken_external_links": broken_external_links,
        "broken_external_links_count": len(broken_external_links),
        "deleted_defined_names": deleted_defined_names,
        "deleted_defined_names_count": len(deleted_defined_names),
        "warnings": warnings,
        "errors": errors,
    }


def convert_with_openpyxl(
    model_path: Path,
    value_path: Path,
    keep_hidden_sheets: bool,
    progress_callback=None,
) -> dict:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ProcessingError(f"缺少 openpyxl 依赖：{exc}") from exc

    keep_vba = model_path.suffix.lower() == ".xlsm"
    formula_wb = None
    data_wb = None
    value_conversion_started = time.time()
    try:
        formula_wb = load_workbook(model_path, data_only=False, read_only=False, keep_vba=keep_vba)
        data_wb = load_workbook(model_path, data_only=True, read_only=False, keep_vba=keep_vba)
        sheet_infos = [
            {"name": sheet.title, "visible": sheet.sheet_state == "visible", "state": sheet.sheet_state or "visible"}
            for sheet in formula_wb.worksheets
        ]
        delete_sheet_names = choose_sheets_to_delete(sheet_infos, keep_hidden_sheets)
        deleted_sheets = []
        for sheet_name in delete_sheet_names:
            if sheet_name in formula_wb.sheetnames:
                del formula_wb[sheet_name]
                deleted_sheets.append(sheet_name)

        formulas_detected_count = 0
        formulas_converted_count = 0
        missing_cached_formulas = []
        processed_sheets = []
        total_sheets = len(formula_wb.worksheets)
        for index, sheet in enumerate(formula_wb.worksheets, start=1):
            processed_sheets.append(sheet.title)
            report_progress(progress_callback, f"正在用缓存值转值工作表：{sheet.title}", index, total_sheets)
            data_sheet = data_wb[sheet.title] if sheet.title in data_wb.sheetnames else None
            for row in sheet.iter_rows():
                for cell in row:
                    if is_formula(cell.value):
                        formulas_detected_count += 1
                        cached = data_sheet[cell.coordinate].value if data_sheet is not None else None
                        if cached is None:
                            missing_cached_formulas.append(f"{sheet.title}!{cell.coordinate}")
                            continue
                        cell.value = cached
                        formulas_converted_count += 1

        broken_external_links = clear_openpyxl_external_links(formula_wb)
        deleted_defined_names = delete_openpyxl_defined_names(formula_wb)
        ensure_openpyxl_visible_sheet(formula_wb)
        formula_wb.save(value_path)
        warnings = [
            "未执行本机重算，仅完成静态扫描和缓存值扫描。",
            "openpyxl 兜底模式可能无法保留形状、控件、部分图表和宏；建议安装或修复 Excel/WPS 自动化后重新生成。",
        ]
        if missing_cached_formulas:
            warnings.append("部分公式没有可读取缓存值，已保留公式并在复检中拦截，不建议交付。")
        return {
            "total_sheets_after_cleanup": len(formula_wb.worksheets),
            "processed_sheets": processed_sheets,
            "processed_sheets_count": len(processed_sheets),
            "deleted_sheets": deleted_sheets,
            "deleted_sheets_count": len(deleted_sheets),
            "skipped_delete_sheets": [],
            "formulas_detected_count": formulas_detected_count,
            "formulas_converted_count": formulas_converted_count,
            "formula_removed": formulas_detected_count == formulas_converted_count,
            "missing_cached_formulas": missing_cached_formulas[:200],
            "missing_cached_formulas_count": len(missing_cached_formulas),
            "broken_external_links": broken_external_links,
            "broken_external_links_count": len(broken_external_links),
            "deleted_defined_names": deleted_defined_names,
            "deleted_defined_names_count": len(deleted_defined_names),
            "processing_engine": "openpyxl",
            "formula_conversion_engine": "openpyxl cached values",
            "recalculation_seconds": 0,
            "value_conversion_seconds": round(time.time() - value_conversion_started, 2),
            "recalculation": {
                "attempted": False,
                "supported": False,
                "message": "未执行本机重算，仅完成静态扫描和缓存值扫描。",
            },
            "warnings": warnings,
            "errors": [],
        }
    except Exception as exc:
        raise ProcessingError(f"openpyxl 生成客户交付值版失败：{exc}") from exc
    finally:
        if formula_wb is not None:
            formula_wb.close()
        if data_wb is not None:
            data_wb.close()


def read_com_sheet_infos(workbook) -> list[dict]:
    infos = []
    for index in range(1, int(workbook.Worksheets.Count) + 1):
        sheet = workbook.Worksheets(index)
        visible_value = int(sheet.Visible)
        state = "visible" if visible_value == XL_SHEET_VISIBLE else "hidden"
        if visible_value == 2:
            state = "veryHidden"
        infos.append({"name": str(sheet.Name), "visible": visible_value == XL_SHEET_VISIBLE, "state": state})
    return infos


def choose_sheets_to_delete(sheet_infos: list[dict], keep_hidden_sheets: bool) -> list[str]:
    candidates = []
    visible_sheet_names = [info["name"] for info in sheet_infos if info.get("visible")]
    for info in sheet_infos:
        name = info["name"]
        should_delete = is_internal_sheet_name(name)
        if not keep_hidden_sheets and not info.get("visible"):
            should_delete = True
        if should_delete:
            candidates.append(name)
    if not sheet_infos:
        return []
    remaining_visible = [name for name in visible_sheet_names if name not in candidates]
    if remaining_visible:
        return candidates
    keep_name = visible_sheet_names[0] if visible_sheet_names else sheet_infos[0]["name"]
    return [name for name in candidates if name != keep_name]


def is_internal_sheet_name(name: str) -> bool:
    return any(keyword in str(name) for keyword in INTERNAL_SHEET_KEYWORDS)


def try_unprotect_workbook(workbook, password: str):
    for candidate in (password, "", "123456"):
        try:
            workbook.Unprotect(Password=candidate)
            return
        except Exception:
            pass


def try_unprotect_sheet(sheet, password: str):
    for candidate in (password, "", "123456"):
        try:
            sheet.Unprotect(Password=candidate)
            return
        except Exception:
            pass


def break_com_external_links(workbook) -> list[dict]:
    broken = []
    for link_type in (1, 2):
        try:
            links = workbook.LinkSources(Type=link_type)
        except Exception:
            links = None
        for link in com_iterable(links):
            try:
                workbook.BreakLink(Name=str(link), Type=link_type)
                broken.append({"type": link_type, "target": str(link)})
            except Exception as exc:
                broken.append({"type": link_type, "target": str(link), "error": str(exc)})
    return broken


def delete_com_defined_names(workbook) -> list[dict]:
    deleted = []
    try:
        count = int(workbook.Names.Count)
    except Exception:
        return deleted
    for index in range(count, 0, -1):
        try:
            defined_name = workbook.Names(index)
            name = str(defined_name.Name)
            refers_to = str(defined_name.RefersTo)
            if should_delete_defined_name(name, refers_to):
                defined_name.Delete()
                deleted.append({"name": name, "reference": refers_to})
        except Exception as exc:
            deleted.append({"name": "", "reference": "", "error": str(exc)})
    return deleted


def clear_openpyxl_external_links(workbook) -> list[dict]:
    links = []
    for link in getattr(workbook, "_external_links", []) or []:
        links.append({"target": str(link)})
    try:
        workbook._external_links = []
    except Exception:
        pass
    return links


def delete_openpyxl_defined_names(workbook) -> list[dict]:
    deleted = []
    defined_names = getattr(workbook, "defined_names", None)
    if not defined_names:
        return deleted
    try:
        items = list(defined_names.items())
    except Exception:
        items = []
    for name, defined_name in items:
        reference = str(getattr(defined_name, "attr_text", "") or "")
        if not should_delete_defined_name(name, reference):
            continue
        try:
            del defined_names[name]
        except Exception:
            try:
                defined_names.pop(name)
            except Exception:
                pass
        deleted.append({"name": name, "reference": reference})
    return deleted


def should_delete_defined_name(name: str, reference: str) -> bool:
    text = f"{name} {reference}"
    lowered = text.lower()
    if str(name).strip() == "_1":
        return True
    if "[" in reference or "http://" in lowered or "https://" in lowered:
        return True
    if reference.lower().startswith("='http") or reference.lower().startswith("='file"):
        return True
    if any(keyword.lower() in lowered for keyword in DEFINED_NAME_BAD_KEYWORDS):
        return True
    if "�" in text or "???" in text:
        return True
    return False


def ensure_openpyxl_visible_sheet(workbook) -> None:
    if any(sheet.sheet_state == "visible" for sheet in workbook.worksheets):
        return
    if workbook.worksheets:
        workbook.worksheets[0].sheet_state = "visible"


def com_iterable(value) -> list:
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return list(value)
    try:
        return list(value)
    except TypeError:
        return [value]
    except Exception:
        return []


def is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def build_delivery_strict_result(delivery_check: dict) -> dict:
    blocking = []
    if delivery_check.get("formula_cells_count", 0) != 0:
        blocking.append({"item": "公式单元格", "count": delivery_check.get("formula_cells_count", 0), "required": 0})
    if delivery_check.get("total_error_count", 0) != 0:
        blocking.append({"item": "公式错误值", "count": delivery_check.get("total_error_count", 0), "required": 0})
    if delivery_check.get("hidden_sheet_count", 0) != 0:
        blocking.append({"item": "隐藏/veryHidden 工作表", "count": delivery_check.get("hidden_sheet_count", 0), "required": 0})
    if delivery_check.get("external_links"):
        blocking.append({"item": "外部链接", "count": len(delivery_check.get("external_links", [])), "required": 0})
    if delivery_check.get("defined_name_issues"):
        blocking.append({"item": "无效命名区域", "count": len(delivery_check.get("defined_name_issues", [])), "required": 0})
    if delivery_check.get("internal_keyword_hits"):
        blocking.append({"item": "旧口径/内部痕迹关键词", "count": len(delivery_check.get("internal_keyword_hits", [])), "required": 0})
    return {"passed": not blocking, "blocking_issues": blocking}


def render_value_version_report(report: dict) -> str:
    source_check = report["source_model_check"]
    delivery_check = report["delivery_value_check"]
    summary = report.get("generation_summary", {})
    lines = [
        "# 财务 Excel 模型版与客户交付值版生成报告",
        "",
        f"- 源文件路径：`{report['source_file_path']}`",
        f"- 源文件识别：{report.get('source_file_role', '-')}",
        f"- 是否误用客户交付值版作为源文件：{'是' if report.get('is_value_version_used_as_source') else '否'}",
        f"- 模型版路径：`{report['model_version_path']}`",
        f"- 客户交付值版路径：`{report['delivery_value_path']}`",
        f"- 最终保存位置：`{report.get('actual_save_location', report['delivery_value_path'])}`",
        f"- 输出模式：{report.get('output_mode', '-')}",
        f"- 检查时间：{report['created_at']}",
        f"- 处理引擎：{summary.get('processing_engine', '-')}",
        f"- 重算状态：{summary.get('recalculation', {}).get('message', '未执行本机重算，仅完成静态扫描和缓存值扫描。')}",
        f"- 最终结论：{report['final_conclusion']}",
        "",
        "## 耗时统计",
        "",
        f"- 源文件检查耗时：{report.get('timing_summary', {}).get('source_file_check_seconds', 0)} 秒",
        f"- Excel COM 重算耗时：{report.get('timing_summary', {}).get('excel_com_recalculation_seconds', 0)} 秒",
        f"- 转值耗时：{report.get('timing_summary', {}).get('value_conversion_seconds', 0)} 秒",
        f"- 复检耗时：{report.get('timing_summary', {}).get('delivery_recheck_seconds', 0)} 秒",
        f"- 总耗时：{report.get('timing_summary', {}).get('total_seconds', report.get('duration_seconds', 0))} 秒",
        "",
        "## 生成结果",
        "",
        f"- 删除内部/隐藏表数量：{summary.get('deleted_sheets_count', 0)}",
        f"- 转值公式数量：{summary.get('formulas_converted_count', 0)} / {summary.get('formulas_detected_count', 0)}",
        f"- 清理外部链接数量：{summary.get('broken_external_links_count', 0)}",
        f"- 清理命名区域数量：{summary.get('deleted_defined_names_count', 0)}",
        f"- 是否可用于交付：{'是' if report['delivery_ready'] else '否'}",
        "",
    ]
    if summary.get("deleted_sheets"):
        lines.append("已删除工作表：" + "、".join(f"`{name}`" for name in summary["deleted_sheets"]))
        lines.append("")
    if report.get("warnings"):
        lines.extend(["## 提示", ""])
        lines.extend(f"- {warning}" for warning in report["warnings"])
        lines.append("")
    if report.get("blocking_issues"):
        lines.extend(["## 阻断项", ""])
        for issue in report["blocking_issues"]:
            lines.append(f"- {issue['item']}：{issue['count']}，要求：{issue['required']}")
        lines.append("")

    lines.extend(["## 源模型版检查结果", ""])
    lines.extend(render_check_brief(source_check).splitlines())
    lines.extend(["", "## 客户交付值版复检结果", ""])
    lines.extend(render_check_brief(delivery_check).splitlines())
    lines.extend(["", "## 客户交付值版详细质检", ""])
    lines.extend(render_markdown_report(delivery_check).splitlines()[1:])
    return "\n".join(lines).rstrip() + "\n"


def render_check_brief(check: dict) -> str:
    other_errors = max(0, int(check.get("total_error_count", 0)) - int(check.get("error_counts", {}).get("#NAME?", 0)))
    lines = [
        f"- 检查文件：`{check.get('checked_file_path', '')}`",
        f"- 工作表总数：{check.get('worksheet_total_count', 0)}",
        f"- 可见表数量：{check.get('visible_sheet_count', 0)}",
        f"- 隐藏表数量：{check.get('hidden_sheet_count', 0)}（veryHidden：{check.get('very_hidden_sheet_count', 0)}）",
        f"- 公式单元格数量：{check.get('formula_cells_count', 0)}",
        f"- #NAME? 数量：{check.get('error_counts', {}).get('#NAME?', 0)}",
        f"- 其他公式错误数量：{other_errors}",
        f"- 内部痕迹关键词命中：{len(check.get('internal_keyword_hits', []))}",
        f"- 外部链接：{len(check.get('external_links', []))}",
        f"- 命名区域异常：{len(check.get('defined_name_issues', []))}",
    ]
    sheet_counts = Counter()
    for location in check.get("error_locations", []):
        if location.get("error_value") == "#NAME?":
            sheet_counts[location.get("sheet_name", "")] += 1
    if sheet_counts:
        lines.append("- 各工作表 #NAME? 数量：" + "；".join(f"{sheet}={count}" for sheet, count in sheet_counts.items()))
    if check.get("total_error_count", 0):
        lines.append("- 各类错误数量：" + "；".join(f"{error}={check.get('error_counts', {}).get(error, 0)}" for error in EXCEL_ERROR_VALUES))
    return "\n".join(lines)
