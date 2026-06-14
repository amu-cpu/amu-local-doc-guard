from __future__ import annotations

import json
import os
import shutil
import subprocess
import time
import warnings
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from uuid import uuid4
from zipfile import ZipFile

from PIL import Image, ImageDraw, ImageFont, ImageGrab

from .processor import OUTPUT_ROOT, ProcessingError, sanitize_filename, write_json


warnings.filterwarnings("ignore", "Palette images with Transparency expressed in bytes.*", UserWarning)

ALLOWED_EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xls"}
PREVIEW_PREFIX = "预览版_"
TEMP_PREVIEW_DIRNAME = "temp_preview"
XL_SHEET_VISIBLE = -1
XL_CELL_TYPE_FORMULAS = -4123
XL_PASTE_VALUES = -4163
XL_SCREEN = 1
XL_PICTURE = -4147
XL_BITMAP = 2
XL_TYPE_PDF = 0
MSO_TEXT_ORIENTATION_HORIZONTAL = 1
MSO_FALSE = 0
MSO_TRUE = -1
SPACING_FACTORS = {
    "compact": (0.95, 95),
    "medium": (1.35, 145),
    "wide": (1.85, 210),
}
OFFICE_PROCESS_NAMES = {"excel", "et", "wps", "wpp"}
REAL_SCREENSHOT_FAILURE_MESSAGE = (
    "真实截图失败，未生成图片化防复制版 Excel。当前环境无法通过 Excel/WPS COM 获取工作表原样截图。"
    "请尝试：1）关闭所有 WPS/Excel 进程后重试；2）用 WPS/Excel 打开原文件并另存为 xlsx；"
    "3）重新启动本工具；4）检查 pywin32 和 WPS/Excel COM 是否可用。"
)
APPROXIMATE_FALLBACK_WARNING = (
    "当前不是原样截图，仅为近似绘制，复杂图表、图片、形状和排版可能丢失，不建议作为客户交付预览版。"
)


@dataclass
class ExcelPreviewOptions:
    watermark_text: str = "保密文件"
    watermark_font_size: int = 28
    watermark_color: str = "#B8B8B8"
    watermark_opacity: int = 20
    watermark_rotation: int = -30
    watermark_spacing: str = "medium"
    protection_password: str = "123456"
    convert_formulas: bool = True
    add_watermark: bool = True
    protect_sheets: bool = True
    protect_workbook_structure: bool = True
    include_hidden_sheets: bool = False
    preview_security_mode: str = "image_based"
    image_range_type: str = "auto"
    allow_approximate_fallback: bool = False
    output_mode: str = "desktop"
    output_report_to_desktop: bool = False
    screenshot_window_mode: str = "quiet"


def process_excel_preview(
    original_name: str,
    save_func,
    output_root: Path = OUTPUT_ROOT,
    options: ExcelPreviewOptions | None = None,
    progress_callback=None,
) -> dict:
    started = time.time()
    options = options or ExcelPreviewOptions()
    original_name = sanitize_filename(original_name)
    suffix = Path(original_name).suffix.lower()
    if suffix not in ALLOWED_EXCEL_EXTENSIONS:
        raise ProcessingError("仅支持 .xlsx、.xlsm 文件；.xls 请先另存为 .xlsx 后处理。")
    if suffix == ".xls":
        raise ProcessingError(".xls 格式请先用 Excel/WPS 另存为 .xlsx 后再处理。")

    safe_stem = sanitize_filename(Path(original_name).stem) or "workbook"
    output_stem = prefixed_preview_stem(safe_stem)
    job_id = f"excel_{safe_stem}_{uuid4().hex[:8]}"
    job_dir = (output_root / job_id).resolve()
    input_dir = job_dir / "input"
    input_dir.mkdir(parents=True, exist_ok=False)

    input_path = input_dir / original_name
    save_func(input_path)

    output_suffix = ".xlsx" if options.preview_security_mode == "image_based" else suffix
    desktop_dir = find_windows_desktop() if options.output_mode == "desktop" else None
    output_version_number = choose_output_version(output_root, desktop_dir, output_stem, output_suffix)
    versioned_stem = versioned_output_stem(output_stem, output_version_number)
    output_excel = job_dir / f"{versioned_stem}{output_suffix}"
    report_path = job_dir / f"{versioned_stem}_处理报告.json"

    warnings: list[str] = []
    errors: list[str] = []
    if options.preview_security_mode == "image_based":
        report = process_image_based_excel(input_path, output_excel, job_dir / "temp_images", options, progress_callback=progress_callback)
    else:
        shutil.copy2(input_path, output_excel)
        try:
            report = process_with_com(output_excel, options, prefer_wps=False, progress_callback=progress_callback)
        except ProcessingError as exc:
            warnings.append(str(exc))
            try:
                report = process_with_com(output_excel, options, prefer_wps=True, progress_callback=progress_callback)
            except ProcessingError as wps_exc:
                warnings.append(str(wps_exc))
                report = process_with_openpyxl(output_excel, options, progress_callback=progress_callback)

    report.update(
        {
            "job_id": job_id,
            "source_file": original_name,
            "output_excel": output_excel.name,
            "output_filename": output_excel.name,
            "report_filename": report_path.name,
            "output_version": f"v{output_version_number}",
            "overwrite_existing": False,
            "excel_download_url": f"/outputs/{job_id}/{output_excel.name}",
            "report_download_url": f"/outputs/{job_id}/{report_path.name}",
            "outputs": {
                "excel": output_excel.name,
                "report": report_path.name,
            },
            "open_password_enabled": False,
            "warnings": dedupe_texts(warnings + report.get("warnings", [])),
            "errors": errors + report.get("errors", []),
            "duration_seconds": round(time.time() - started, 2),
            "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
            "local_only": True,
        }
    )
    prepare_excel_output_location(report, output_excel, report_path, options, desktop_dir=desktop_dir)
    write_json(report_path, report)
    finalize_report_output_copy(report, report_path)
    write_json(report_path, report)
    return report


def inspect_excel_upload(
    original_name: str,
    save_func,
    output_root: Path = OUTPUT_ROOT,
) -> dict:
    original_name = sanitize_filename(original_name)
    suffix = Path(original_name).suffix.lower()
    if suffix not in ALLOWED_EXCEL_EXTENSIONS:
        raise ProcessingError("仅支持 .xlsx、.xlsm 文件；.xls 请先另存为 .xlsx 后处理。")
    if suffix == ".xls":
        raise ProcessingError(".xls 格式请先用 Excel/WPS 另存为 .xlsx 后再处理。")

    safe_stem = sanitize_filename(Path(original_name).stem) or "workbook"
    upload_id = f"excel_preview_{safe_stem}_{uuid4().hex[:8]}"
    upload_dir = (output_root / TEMP_PREVIEW_DIRNAME / upload_id).resolve()
    upload_dir.mkdir(parents=True, exist_ok=False)
    source_path = upload_dir / original_name
    save_func(source_path)
    sheets = list_workbook_sheets(source_path)
    visible_sheets = [sheet for sheet in sheets if sheet["visible"]]
    meta = {
        "upload_id": upload_id,
        "original_name": original_name,
        "source_file": source_path.name,
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "sheets": sheets,
        "default_sheet": (visible_sheets[0] if visible_sheets else sheets[0])["name"] if sheets else "",
    }
    write_json(upload_dir / "upload_meta.json", meta)
    return meta


def process_excel_preview_from_upload(
    upload_id: str,
    output_root: Path = OUTPUT_ROOT,
    options: ExcelPreviewOptions | None = None,
    progress_callback=None,
) -> dict:
    upload_meta = read_upload_meta(upload_id, output_root)
    source_path = upload_source_path(upload_meta, output_root)

    def save_func(dest):
        shutil.copy2(source_path, dest)

    return process_excel_preview(upload_meta["original_name"], save_func, output_root, options=options, progress_callback=progress_callback)


def prepare_excel_output_location(report: dict, output_excel: Path, report_path: Path, options: ExcelPreviewOptions, desktop_dir: Path | None = None):
    app_output_path = Path(output_excel).resolve()
    app_report_path = Path(report_path).resolve()
    report.update(
        {
            "output_mode": options.output_mode,
            "excel_output_path": str(app_output_path),
            "report_output_path": str(app_report_path),
            "excel_desktop_path": None,
            "report_desktop_path": None,
            "desktop_excel_copied": False,
            "desktop_report_copied": False,
            "app_output_path": str(app_output_path),
            "app_report_path": str(app_report_path),
            "desktop_output_enabled": options.output_mode == "desktop",
            "desktop_output_path": "",
            "desktop_report_path": None,
            "actual_save_location": str(app_output_path),
            "report_location": "程序 output 目录，可点击“下载处理报告 JSON”查看。",
            "output_copy_status": "app_output_only",
        }
    )
    if options.output_mode != "desktop":
        return

    desktop_dir = desktop_dir or find_windows_desktop()
    if not desktop_dir:
        report["output_copy_status"] = "desktop_not_found_fallback_app_output"
        report.setdefault("warnings", []).append(f"桌面输出失败，已保存在程序 output 目录：{app_output_path}")
        return

    try:
        desktop_dir.mkdir(parents=True, exist_ok=True)
        desktop_excel = desktop_dir / output_excel.name
        if desktop_excel.exists():
            raise ProcessingError(f"桌面已存在同名文件：{desktop_excel}")
        shutil.copy2(app_output_path, desktop_excel)
        report["excel_desktop_path"] = str(desktop_excel)
        report["desktop_output_path"] = str(desktop_excel)
        report["desktop_excel_copied"] = True
        if options.output_report_to_desktop:
            report["report_desktop_path"] = str(desktop_dir / report_path.name)
            report["desktop_report_path"] = str(desktop_dir / report_path.name)
        report["actual_save_location"] = str(desktop_excel)
        report["output_copy_status"] = "desktop_excel_copied"
        report["desktop_output_message"] = f"已输出到桌面：{output_excel.name}"
    except Exception as exc:
        report["output_copy_status"] = "desktop_excel_copy_failed_fallback_app_output"
        report.setdefault("warnings", []).append(f"桌面输出失败，已保存在程序 output 目录：{app_output_path}。原因：{exc}")


def finalize_report_output_copy(report: dict, report_path: Path):
    desktop_report_path = report.get("report_desktop_path")
    if not desktop_report_path:
        return
    try:
        desktop_report = Path(desktop_report_path)
        if desktop_report.exists():
            raise ProcessingError(f"桌面已存在同名报告：{desktop_report}")
        report["output_copy_status"] = "desktop_excel_and_report_copied"
        report["desktop_report_copied"] = True
        write_json(report_path, report)
        shutil.copy2(report_path, desktop_report)
    except Exception as exc:
        report["desktop_report_copied"] = False
        report["output_copy_status"] = "desktop_report_copy_failed_excel_copied"
        report.setdefault("warnings", []).append(f"处理报告复制到桌面失败，程序 output 目录中仍保留报告。原因：{exc}")


def find_windows_desktop() -> Path | None:
    candidates = []
    userprofile = os.environ.get("USERPROFILE")
    if userprofile:
        candidates.append(Path(userprofile) / "Desktop")
    onedrive = os.environ.get("OneDrive") or os.environ.get("OneDriveConsumer")
    if onedrive:
        candidates.append(Path(onedrive) / "Desktop")
    candidates.append(Path.home() / "Desktop")
    for candidate in candidates:
        try:
            if candidate.exists() and candidate.is_dir():
                return candidate.resolve()
        except OSError:
            continue
    return candidates[0].resolve() if candidates else None


def choose_output_version(output_root: Path, desktop_dir: Path | None, output_stem: str, output_suffix: str) -> int:
    output_root = Path(output_root).resolve()
    for version in range(1, 1000):
        stem = versioned_output_stem(output_stem, version)
        excel_name = f"{stem}{output_suffix}"
        report_name = f"{stem}_处理报告.json"
        if output_name_exists(output_root, excel_name) or output_name_exists(output_root, report_name):
            continue
        if desktop_dir and ((desktop_dir / excel_name).exists() or (desktop_dir / report_name).exists()):
            continue
        return version
    raise ProcessingError("无法生成唯一版本文件名，请清理旧输出后重试。")


def versioned_output_stem(output_stem: str, version: int) -> str:
    return output_stem if version <= 1 else f"{output_stem}_v{version}"


def output_name_exists(output_root: Path, filename: str) -> bool:
    if not output_root.exists():
        return False
    for path in output_root.rglob(filename):
        if path.is_file():
            return True
    return False


def generate_watermark_preview(
    upload_id: str,
    sheet_name: str,
    range_type: str,
    options: ExcelPreviewOptions,
    output_root: Path = OUTPUT_ROOT,
) -> dict:
    started = time.time()
    upload_meta = read_upload_meta(upload_id, output_root)
    source_path = upload_source_path(upload_meta, output_root)
    upload_dir = source_path.parent
    preview_name = f"watermark_preview_{uuid4().hex[:8]}.png"
    preview_path = (upload_dir / preview_name).resolve()
    result = capture_sheet_to_png(
        source_path,
        sheet_name,
        range_type,
        preview_path,
        options,
        temp_dir=upload_dir,
        allow_fallback=options.allow_approximate_fallback,
        prefer_silent=False,
    )
    if not preview_path.exists() or preview_path.stat().st_size == 0:
        raise ProcessingError("预览图片生成失败，请重新生成预览图。")
    if result.get("warnings"):
        deduped_warnings = []
        for warning in result["warnings"]:
            if warning and warning not in deduped_warnings:
                deduped_warnings.append(warning)
        result["warnings"] = deduped_warnings
    result.update(
        {
            "upload_id": upload_id,
            "sheet_name": sheet_name,
            "preview_file": f"{TEMP_PREVIEW_DIRNAME}/{upload_id}/{preview_name}",
            "duration_seconds": round(time.time() - started, 2),
        }
    )
    return result


def process_image_based_excel(source_path: Path, output_excel: Path, temp_dir: Path, options: ExcelPreviewOptions, progress_callback=None) -> dict:
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as OpenpyxlImage
    except Exception as exc:
        raise ProcessingError(f"缺少 openpyxl 依赖，无法生成图片化 Excel：{exc}") from exc

    temp_dir.mkdir(parents=True, exist_ok=True)
    sheets = list_workbook_sheets(source_path)
    processed_sheets = []
    skipped_hidden_sheets = []
    warnings_list: list[str] = []
    errors: list[str] = []
    engines: set[str] = set()
    capture_methods: set[str] = set()
    image_based_sheets_count = 0
    watermark_sheets_count = 0
    protected_sheets_count = 0

    target_sheets = []
    for sheet in sheets:
        if not sheet["visible"] and not options.include_hidden_sheets:
            skipped_hidden_sheets.append(sheet["name"])
            continue
        target_sheets.append(sheet)
    has_visible_target = any(sheet["visible"] for sheet in target_sheets)

    formulas_detected_count = count_source_formulas(source_path, {sheet["name"] for sheet in target_sheets})

    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    image_paths: dict[str, Path] = {}
    for index, sheet_info in enumerate(target_sheets, start=1):
        image_paths[sheet_info["name"]] = temp_dir / f"sheet_{index:03d}_{uuid4().hex[:8]}.png"

    capture_results = {}
    if target_sheets:
        capture_results = capture_workbook_sheets_to_pngs(
            source_path,
            [sheet["name"] for sheet in target_sheets],
            options.image_range_type,
            image_paths,
            options,
            temp_dir,
            allow_fallback=options.allow_approximate_fallback,
            window_mode=options.screenshot_window_mode,
            progress_callback=progress_callback,
        )

    for index, sheet_info in enumerate(target_sheets, start=1):
        sheet_name = sheet_info["name"]
        processed_sheets.append(sheet_name)
        output_sheet = workbook.create_sheet(title=sheet_name)
        output_sheet.sheet_view.showGridLines = False
        try:
            image_path = image_paths[sheet_name]
            image_result = capture_results.get(sheet_name)
            if not image_result:
                raise ProcessingError("当前工作表没有生成截图。")
            engines.add(image_result.get("preview_engine", "openpyxl + Pillow"))
            capture_methods.add(image_result.get("capture_method", "fallback_openpyxl"))
            warnings_list.extend(image_result.get("warnings", []))
            report_progress(progress_callback, f"正在写入截图到 Excel：{sheet_name}", index, len(target_sheets))
            image = OpenpyxlImage(str(image_path))
            output_sheet.add_image(image, "A1")
            try:
                output_sheet.column_dimensions["A"].width = min(255, max(12, image.width / 7))
                output_sheet.row_dimensions[1].height = min(409, max(80, image.height * 0.75))
            except Exception:
                pass
            image_based_sheets_count += 1
            if options.add_watermark:
                watermark_sheets_count += 1
        except Exception as exc:
            raise ProcessingError(f"工作表“{sheet_name}”真实截图失败：{exc}") from exc

        if options.protect_sheets:
            protect_openpyxl_sheet(output_sheet, options.protection_password)
            protected_sheets_count += 1

        if has_visible_target and sheet_info.get("state") in {"hidden", "veryHidden"} and len(target_sheets) > 1:
            output_sheet.sheet_state = sheet_info["state"]

    if not target_sheets:
        output_sheet = workbook.create_sheet(title="预览")
        output_sheet["A1"] = "没有可处理的工作表。"
        processed_sheets.append("预览")
        protect_openpyxl_sheet(output_sheet, options.protection_password)
        protected_sheets_count = 1

    workbook_structure_protected = False
    if options.protect_workbook_structure:
        report_progress(progress_callback, "正在保护工作簿结构...", len(target_sheets), len(target_sheets))
        workbook.security.lockStructure = True
        workbook.security.workbookPassword = options.protection_password
        workbook_structure_protected = True

    report_progress(progress_callback, "正在写入预览版 Excel...", len(target_sheets), len(target_sheets))
    workbook.save(output_excel)
    report_progress(progress_callback, "正在检查图片插入结果...", len(target_sheets), len(target_sheets))
    image_insert_check = validate_image_based_output(output_excel, processed_sheets)
    if not image_insert_check["image_insert_check_passed"]:
        missing = "、".join(image_insert_check["sheets_missing_images"]) or "未知工作表"
        raise ProcessingError(f"图片化 Excel 导出自检失败，以下工作表没有插入截图图片：{missing}")
    shutil.rmtree(temp_dir, ignore_errors=True)

    engine = " / ".join(sorted(engines)) if engines else "openpyxl + Pillow"
    screenshot_engine = screenshot_engine_from_engines(engines)
    deduped_warnings = dedupe_texts(warnings_list)
    if screenshot_engine == "fallback_openpyxl" or "openpyxl + Pillow" in engines:
        deduped_warnings.append(APPROXIMATE_FALLBACK_WARNING)
        deduped_warnings = dedupe_texts(deduped_warnings)

    visual_objects_preserved = True if screenshot_engine in {"Excel COM", "WPS COM"} and "openpyxl + Pillow" not in engines else "unknown"
    used_approximate_fallback = screenshot_engine == "fallback_openpyxl" or "openpyxl + Pillow" in engines
    return {
        "preview_security_mode": "image_based",
        "preview_security_mode_label": "图片化防复制版 Excel",
        "total_sheets": len(sheets),
        "processed_sheets": processed_sheets,
        "skipped_hidden_sheets": skipped_hidden_sheets,
        "processed_sheets_count": len(processed_sheets),
        "skipped_hidden_sheets_count": len(skipped_hidden_sheets),
        "image_based_sheets_count": image_based_sheets_count,
        "formulas_detected_count": formulas_detected_count,
        "formulas_converted_count": 0,
        "formula_conversion_required": False,
        "formula_conversion_status": "skipped_image_based",
        "formula_conversion_engine": "skipped_image_based",
        "formula_conversion_note": "公式转数值：已跳过（图片化模式无需转值）。",
        "watermark_enabled": options.add_watermark,
        "watermark_text": options.watermark_text,
        "watermark_style": watermark_style_dict(options),
        "watermark_sheets_count": watermark_sheets_count,
        "protected_sheets_count": protected_sheets_count,
        "workbook_structure_protected": workbook_structure_protected,
        "open_password_enabled": False,
        "edit_protection_enabled": options.protect_sheets or options.protect_workbook_structure,
        "real_cell_data_removed": True,
        "formula_removed": True,
        "copy_risk_level": "high" if used_approximate_fallback else "low",
        "screenshot_engine": screenshot_engine,
        "window_mode": options.screenshot_window_mode,
        "screenshot_capture_methods": sorted(capture_methods),
        "images_preserved_in_screenshot": visual_objects_preserved,
        "charts_preserved_in_screenshot": visual_objects_preserved,
        "image_insert_check_passed": image_insert_check["image_insert_check_passed"],
        "inserted_images_count": image_insert_check["inserted_images_count"],
        "expected_sheets_count": image_insert_check["expected_sheets_count"],
        "sheets_missing_images": image_insert_check["sheets_missing_images"],
        "xlsx_media_count": image_insert_check["xlsx_media_count"],
        "xlsx_drawings_count": image_insert_check["xlsx_drawings_count"],
        "preview_and_export_same_engine": True,
        "screenshot_function_shared": True,
        "risk_notice": (
            "当前为近似绘制，不建议交付客户。"
            if used_approximate_fallback
            else "该文件已尽量转为图片化预览，不能直接复制单元格数据，但仍无法防止截图、拍照或 OCR 识别。"
        ),
        "processing_engine": engine,
        "warnings": dedupe_texts(
            ["图片化防复制版未执行公式转数值，因为最终文件只保留图片化预览，不保留原始公式和单元格数据。"]
            + deduped_warnings
        ),
        "errors": errors,
    }


def generate_sheet_image(
    workbook_path: Path,
    sheet_name: str,
    range_type: str,
    options: ExcelPreviewOptions,
    image_path: Path,
    temp_dir: Path,
    allow_com: bool = True,
    prefer_silent: bool = False,
    allow_fallback: bool = False,
) -> dict:
    return capture_sheet_to_png(
        workbook_path,
        sheet_name,
        range_type,
        image_path,
        options,
        temp_dir=temp_dir,
        allow_fallback=allow_fallback,
        prefer_silent=prefer_silent,
        allow_com=allow_com,
    )


def capture_workbook_sheets_to_pngs(
    workbook_path: Path,
    sheet_names: list[str],
    range_type: str,
    output_paths: dict[str, Path],
    options: ExcelPreviewOptions,
    temp_dir: Path,
    allow_fallback: bool = False,
    window_mode: str = "quiet",
    progress_callback=None,
) -> dict[str, dict]:
    errors: list[str] = []
    for prefer_wps in (False, True):
        engine_name = "WPS COM" if prefer_wps else "Excel COM"
        try:
            return capture_workbook_sheets_with_com(
                workbook_path,
                sheet_names,
                range_type,
                output_paths,
                options,
                temp_dir,
                prefer_wps=prefer_wps,
                window_mode=window_mode,
                progress_callback=progress_callback,
            )
        except ProcessingError as exc:
            errors.append(f"{engine_name}：{exc}")

    if allow_fallback:
        results = {}
        for index, sheet_name in enumerate(sheet_names, start=1):
            report_progress(progress_callback, f"正在近似绘制：{sheet_name}", index, len(sheet_names))
            result = generate_preview_with_openpyxl(workbook_path, sheet_name, range_type, options, output_paths[sheet_name])
            result.setdefault("warnings", []).extend(errors + [APPROXIMATE_FALLBACK_WARNING])
            result["screenshot_function_shared"] = True
            result["warnings"] = dedupe_texts(result.get("warnings", []))
            results[sheet_name] = result
        return results

    hint = "安静模式截图失败，可切换到可见调试模式后重试。" if window_mode == "quiet" else ""
    detail = "；".join(errors)
    raise ProcessingError(f"{REAL_SCREENSHOT_FAILURE_MESSAGE}{hint} 失败详情：{detail}")


def capture_workbook_sheets_with_com(
    workbook_path: Path,
    sheet_names: list[str],
    range_type: str,
    output_paths: dict[str, Path],
    options: ExcelPreviewOptions,
    temp_dir: Path,
    prefer_wps: bool,
    window_mode: str,
    progress_callback=None,
) -> dict[str, dict]:
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        raise ProcessingError(f"Excel/WPS COM 不可用：{exc}") from exc

    engine_name = "WPS COM" if prefer_wps else "Excel COM"
    workbook_path = Path(workbook_path).resolve()
    temp_dir = Path(temp_dir).resolve()
    temp_dir.mkdir(parents=True, exist_ok=True)
    temp_copy = (temp_dir / f"workbook_capture_{uuid4().hex[:8]}{workbook_path.suffix}").resolve()
    shutil.copy2(workbook_path, temp_copy)
    app = None
    workbook = None
    app_process_ids: set[int] = set()
    office_processes_before = office_process_snapshot()
    pythoncom.CoInitialize()
    try:
        app = win32com.client.DispatchEx("Ket.Application" if prefer_wps else "Excel.Application")
        app_process_ids = com_app_process_ids(app)
        configure_com_capture_window(app, window_mode)
        workbook = open_com_workbook(app, temp_copy, read_only=True)
        time.sleep(0.4)
        configure_active_com_window(app, window_mode)
        results = {}
        for index, sheet_name in enumerate(sheet_names, start=1):
            report_progress(progress_callback, f"正在截图：{sheet_name}", index, len(sheet_names))
            output_path = Path(output_paths[sheet_name]).resolve()
            configure_active_com_window(app, window_mode)
            result = capture_open_workbook_sheet_to_png(
                engine_name,
                app,
                workbook,
                sheet_name,
                range_type,
                options,
                output_path,
                window_mode=window_mode,
            )
            result["screenshot_function_shared"] = True
            result["window_mode"] = window_mode
            results[sheet_name] = result
        return results
    except ProcessingError:
        raise
    except Exception as exc:
        message = str(exc)
        if looks_like_locked_file_error(message):
            raise ProcessingError("文件正在被 WPS/Excel 占用，请先关闭后再处理。") from exc
        raise ProcessingError(f"{engine_name} 批量截图失败：{message}") from exc
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if app is not None:
            cleanup_com_application(app, app_process_ids)
        cleanup_new_office_processes(office_processes_before)
        if temp_copy.exists():
            try:
                temp_copy.unlink()
            except OSError:
                pass
        pythoncom.CoUninitialize()


def configure_com_capture_window(app, window_mode: str):
    try:
        app.Visible = True
        app.DisplayAlerts = False
    except Exception:
        pass
    try:
        app.ScreenUpdating = True
        app.EnableEvents = False
    except Exception:
        pass
    if window_mode == "quiet":
        move_com_window_offscreen(app)


def configure_active_com_window(app, window_mode: str):
    if window_mode == "quiet":
        move_com_window_offscreen(app)


def move_com_window_offscreen(app):
    targets = [app]
    try:
        active_window = app.ActiveWindow
    except Exception:
        active_window = None
    if active_window is not None:
        targets.append(active_window)
    for target in targets:
        try:
            target.WindowState = -4143
        except Exception:
            pass
        for attr, value in (("Left", -32000), ("Top", -32000), ("Width", 900), ("Height", 700)):
            try:
                setattr(target, attr, value)
            except Exception:
                pass


def com_app_process_ids(app) -> set[int]:
    process_ids: set[int] = set()
    hwnd_candidates = []
    for attr in ("Hwnd", "HWND", "hWnd"):
        try:
            hwnd = int(getattr(app, attr) or 0)
            if hwnd:
                hwnd_candidates.append(hwnd)
        except Exception:
            pass
    try:
        active_window = app.ActiveWindow
        for attr in ("Hwnd", "HWND", "hWnd"):
            try:
                hwnd = int(getattr(active_window, attr) or 0)
                if hwnd:
                    hwnd_candidates.append(hwnd)
            except Exception:
                pass
    except Exception:
        pass
    if not hwnd_candidates:
        return process_ids
    try:
        import win32process
    except Exception:
        return process_ids
    for hwnd in hwnd_candidates:
        try:
            _thread_id, process_id = win32process.GetWindowThreadProcessId(hwnd)
            if process_id:
                process_ids.add(int(process_id))
        except Exception:
            continue
    return process_ids


def cleanup_com_application(app, process_ids: set[int]):
    try:
        app.DisplayAlerts = False
    except Exception:
        pass
    try:
        app.Quit()
    except Exception:
        pass
    if not process_ids:
        return
    time.sleep(0.6)
    for process_id in process_ids:
        terminate_process_if_running(process_id)


def office_process_snapshot() -> dict[int, str]:
    command = (
        "$names=@('EXCEL','et','wps','wpp');"
        "Get-Process | Where-Object { $names -contains $_.ProcessName } | "
        "Select-Object ProcessName,Id | ConvertTo-Json -Compress"
    )
    try:
        completed = subprocess.run(
            ["powershell", "-NoProfile", "-Command", command],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="ignore",
            timeout=5,
        )
    except Exception:
        return {}
    if completed.returncode != 0:
        return {}
    data = (completed.stdout or "").strip()
    if not data:
        return {}
    try:
        parsed = json.loads(data)
    except Exception:
        return {}
    if isinstance(parsed, dict):
        parsed = [parsed]
    snapshot: dict[int, str] = {}
    parsed_items = parsed if isinstance(parsed, list) else []
    for item in parsed_items:
        try:
            process_id = int(item.get("Id") or 0)
            process_name = str(item.get("ProcessName") or "").lower()
        except Exception:
            continue
        if process_id and process_name in OFFICE_PROCESS_NAMES:
            snapshot[process_id] = process_name
    return snapshot


def cleanup_new_office_processes(before_snapshot: dict[int, str]):
    if before_snapshot is None:
        before_snapshot = {}
    time.sleep(0.8)
    for _attempt in range(2):
        after_snapshot = office_process_snapshot()
        new_process_ids = [
            process_id
            for process_id, process_name in after_snapshot.items()
            if process_id not in before_snapshot and process_name in OFFICE_PROCESS_NAMES
        ]
        if not new_process_ids:
            return
        for process_id in new_process_ids:
            terminate_process_if_running(process_id)
        time.sleep(0.5)


def terminate_process_if_running(process_id: int):
    try:
        import win32api
        import win32con
        import win32process
    except Exception:
        return
    try:
        query_flag = getattr(win32con, "PROCESS_QUERY_LIMITED_INFORMATION", 0x1000)
        handle = win32api.OpenProcess(query_flag | win32con.PROCESS_TERMINATE, False, int(process_id))
    except Exception:
        return
    try:
        exit_code = win32process.GetExitCodeProcess(handle)
        if exit_code == 259:
            win32api.TerminateProcess(handle, 0)
    except Exception:
        pass
    finally:
        try:
            handle.Close()
        except Exception:
            pass


def capture_sheet_to_png(
    workbook_path: Path,
    sheet_name: str,
    range_type: str,
    output_png_path: Path,
    options: ExcelPreviewOptions,
    temp_dir: Path | None = None,
    allow_fallback: bool = False,
    prefer_silent: bool = False,
    allow_com: bool = True,
) -> dict:
    if not allow_com and not allow_fallback:
        raise ProcessingError(REAL_SCREENSHOT_FAILURE_MESSAGE)
    workbook_path = Path(workbook_path).resolve()
    output_png_path = Path(output_png_path).resolve()
    output_png_path.parent.mkdir(parents=True, exist_ok=True)
    temp_dir = Path(temp_dir or output_png_path.parent).resolve()
    temp_dir.mkdir(parents=True, exist_ok=True)
    temp_dir = Path(temp_dir).resolve()
    temp_copy = (temp_dir / f"sheet_capture_{uuid4().hex[:8]}{workbook_path.suffix}").resolve()
    shutil.copy2(workbook_path, temp_copy)
    errors: list[str] = []
    try:
        if allow_com:
            try:
                result = generate_preview_with_com(temp_copy, sheet_name, range_type, options, output_png_path, prefer_wps=False, prefer_silent=prefer_silent)
                ensure_png_created(output_png_path)
                result["screenshot_function_shared"] = True
                if result.get("warnings"):
                    result["warnings"] = dedupe_texts(result["warnings"])
                return result
            except ProcessingError as exc:
                errors.append(f"Excel COM：{exc}")
            try:
                result = generate_preview_with_com(temp_copy, sheet_name, range_type, options, output_png_path, prefer_wps=True, prefer_silent=prefer_silent)
                ensure_png_created(output_png_path)
                result.setdefault("warnings", []).insert(0, "Excel COM 无法处理该文件，已自动切换 WPS COM 真实截图。")
                result["screenshot_function_shared"] = True
                result["warnings"] = dedupe_texts(result.get("warnings", []))
                return result
            except ProcessingError as exc:
                errors.append(f"WPS COM：{exc}")

        if allow_fallback:
            result = generate_preview_with_openpyxl(workbook_path, sheet_name, range_type, options, output_png_path)
            ensure_png_created(output_png_path)
            result.setdefault("warnings", []).extend(errors + [APPROXIMATE_FALLBACK_WARNING])
            result["screenshot_function_shared"] = True
            result["warnings"] = dedupe_texts(result.get("warnings", []))
            return result

        detail = "；".join(errors)
        raise ProcessingError(f"{REAL_SCREENSHOT_FAILURE_MESSAGE} 失败详情：{detail}")
    finally:
        if temp_copy.exists():
            try:
                temp_copy.unlink()
            except OSError:
                pass


def ensure_png_created(path: Path):
    path = Path(path)
    if not path.exists() or path.stat().st_size == 0:
        raise ProcessingError("截图图片生成失败。")


def validate_image_based_output(output_excel: Path, expected_sheet_names: list[str]) -> dict:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ProcessingError(f"缺少 openpyxl，无法检查图片化 Excel 输出：{exc}") from exc

    output_excel = Path(output_excel).resolve()
    if not output_excel.exists() or output_excel.stat().st_size == 0:
        raise ProcessingError("图片化 Excel 输出文件不存在。")

    inserted_images_count = 0
    sheets_missing_images: list[str] = []
    workbook = load_workbook(output_excel, data_only=False)
    try:
        for sheet_name in expected_sheet_names:
            if sheet_name not in workbook.sheetnames:
                sheets_missing_images.append(sheet_name)
                continue
            sheet = workbook[sheet_name]
            image_count = len(getattr(sheet, "_images", []))
            inserted_images_count += image_count
            if image_count < 1:
                sheets_missing_images.append(sheet_name)
    finally:
        workbook.close()

    media_count = 0
    drawings_count = 0
    try:
        with ZipFile(output_excel) as archive:
            names = archive.namelist()
            media_count = len([name for name in names if name.startswith("xl/media/")])
            drawings_count = len([name for name in names if name.startswith("xl/drawings/")])
    except Exception as exc:
        raise ProcessingError(f"无法检查图片化 Excel 内部图片结构：{exc}") from exc

    expected_count = len(expected_sheet_names)
    check_passed = (
        expected_count > 0
        and inserted_images_count >= expected_count
        and not sheets_missing_images
        and media_count > 0
        and drawings_count > 0
    )
    return {
        "image_insert_check_passed": check_passed,
        "inserted_images_count": inserted_images_count,
        "expected_sheets_count": expected_count,
        "sheets_missing_images": sheets_missing_images,
        "xlsx_media_count": media_count,
        "xlsx_drawings_count": drawings_count,
    }


def process_with_com(output_excel: Path, options: ExcelPreviewOptions, prefer_wps: bool, progress_callback=None) -> dict:
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        raise ProcessingError(f"缺少 Excel/WPS 自动化依赖 pywin32：{exc}") from exc

    engine_name = "WPS COM" if prefer_wps else "Excel COM"
    app = None
    workbook = None
    pythoncom.CoInitialize()
    try:
        prog_id = "Ket.Application" if prefer_wps else "Excel.Application"
        app = win32com.client.DispatchEx(prog_id)
        app.Visible = False
        app.DisplayAlerts = False
        try:
            app.ScreenUpdating = False
            app.EnableEvents = False
            app.AskToUpdateLinks = False
        except Exception:
            pass

        workbook = open_com_workbook(app, output_excel, read_only=False)
        if options.convert_formulas:
            calculate_workbook(app)

        summary = process_com_workbook(workbook, app, options, progress_callback=progress_callback)
        workbook.Save()
        summary["processing_engine"] = engine_name
        summary["formula_conversion_engine"] = engine_name if options.convert_formulas else "disabled"
        if prefer_wps and options.protect_sheets:
            summary.setdefault("warnings", []).append("当前环境可能无法完全禁止选择和复制锁定单元格。")
        return summary
    except Exception as exc:
        message = str(exc)
        if looks_like_locked_file_error(message):
            raise ProcessingError("请先关闭正在打开的 Excel 文件后再处理。") from exc
        raise ProcessingError(f"{engine_name} 处理失败：{message}") from exc
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if app is not None:
            try:
                app.DisplayAlerts = False
                app.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def process_com_workbook(workbook, app, options: ExcelPreviewOptions, progress_callback=None) -> dict:
    processed_sheets = []
    skipped_hidden_sheets = []
    formulas_detected_count = 0
    formulas_converted_count = 0
    watermark_sheets_count = 0
    protected_sheets_count = 0
    warnings: list[str] = []
    errors: list[str] = []

    total_sheets = int(workbook.Worksheets.Count)
    for sheet_index in range(1, total_sheets + 1):
        sheet = workbook.Worksheets(sheet_index)
        sheet_name = str(sheet.Name)
        report_progress(progress_callback, f"正在处理工作表：{sheet_name}", sheet_index, total_sheets)
        is_visible = int(sheet.Visible) == XL_SHEET_VISIBLE
        if not is_visible and not options.include_hidden_sheets:
            skipped_hidden_sheets.append(sheet_name)
            continue

        processed_sheets.append(sheet_name)
        try:
            try_unprotect_sheet(sheet, options.protection_password)
            used_range = sheet.UsedRange
            if options.convert_formulas:
                formulas_count = count_formulas(used_range)
                formulas_detected_count += formulas_count
                if formulas_count:
                    used_range.Copy()
                    used_range.PasteSpecial(Paste=XL_PASTE_VALUES)
                    app.CutCopyMode = False
                    formulas_converted_count += formulas_count

            if options.add_watermark:
                add_com_watermarks(sheet, used_range, options)
                watermark_sheets_count += 1

            if options.protect_sheets:
                lock_sheet_cells(sheet)
                protect_sheet(sheet, options.protection_password)
                protected_sheets_count += 1
        except Exception as exc:
            errors.append(f"{sheet_name}: {exc}")

    workbook_structure_protected = False
    if options.protect_workbook_structure:
        try:
            workbook.Protect(Password=options.protection_password, Structure=True, Windows=False)
            workbook_structure_protected = True
        except Exception as exc:
            errors.append(f"工作簿结构保护失败：{exc}")

    return {
        "total_sheets": total_sheets,
        "processed_sheets": processed_sheets,
        "skipped_hidden_sheets": skipped_hidden_sheets,
        "processed_sheets_count": len(processed_sheets),
        "skipped_hidden_sheets_count": len(skipped_hidden_sheets),
        "preview_security_mode": "locked_excel",
        "preview_security_mode_label": "普通锁定版 Excel",
        "image_based_sheets_count": 0,
        "formulas_detected_count": formulas_detected_count,
        "formulas_converted_count": formulas_converted_count,
        "formula_conversion_required": options.convert_formulas,
        "formula_conversion_status": formula_conversion_status(options.convert_formulas, formulas_detected_count, formulas_converted_count, not errors),
        "formula_conversion_note": "未检测到公式单元格。" if formulas_detected_count == 0 else "",
        "watermark_enabled": options.add_watermark,
        "watermark_text": options.watermark_text,
        "watermark_style": watermark_style_dict(options),
        "watermark_sheets_count": watermark_sheets_count,
        "protected_sheets_count": protected_sheets_count,
        "workbook_structure_protected": workbook_structure_protected,
        "edit_protection_enabled": options.protect_sheets or options.protect_workbook_structure,
        "real_cell_data_removed": False,
        "formula_removed": options.convert_formulas and formulas_converted_count == formulas_detected_count,
        "copy_risk_level": "medium",
        "screenshot_engine": "not_used",
        "images_preserved_in_screenshot": "unknown",
        "charts_preserved_in_screenshot": "unknown",
        "risk_notice": "该文件已保护工作表，但 Excel 工作表保护不是强加密，仍可能被复制或绕过。如需减少复制风险，请使用图片化防复制版。",
        "warnings": warnings,
        "errors": errors,
    }


def calculate_workbook(app):
    try:
        app.CalculateFullRebuild()
    except Exception:
        try:
            app.CalculateFull()
        except Exception:
            app.Calculate()


def try_unprotect_sheet(sheet, password: str):
    try:
        sheet.Unprotect(Password=password)
    except Exception:
        pass


def count_formulas(used_range) -> int:
    try:
        return count_formula_values(used_range.Formula)
    except Exception:
        try:
            formulas = used_range.SpecialCells(XL_CELL_TYPE_FORMULAS)
            return int(getattr(formulas, "CountLarge", formulas.Count))
        except Exception:
            return 0


def count_formula_values(value) -> int:
    if isinstance(value, str):
        return 1 if value.startswith("=") else 0
    if isinstance(value, (tuple, list)):
        return sum(count_formula_values(item) for item in value)
    return 0


def lock_sheet_cells(sheet):
    try:
        sheet.Cells.Locked = True
    except Exception:
        pass


def protect_sheet(sheet, password: str):
    sheet.Protect(
        Password=password,
        DrawingObjects=True,
        Contents=True,
        Scenarios=True,
        AllowFiltering=True,
        AllowSorting=False,
        AllowFormattingCells=False,
        AllowFormattingColumns=False,
        AllowFormattingRows=False,
        AllowInsertingColumns=False,
        AllowInsertingRows=False,
        AllowDeletingColumns=False,
        AllowDeletingRows=False,
    )
    try:
        sheet.EnableSelection = 0
    except Exception:
        pass


def add_com_watermarks(sheet, used_range, options: ExcelPreviewOptions):
    left = float(getattr(used_range, "Left", 0) or 0)
    top = float(getattr(used_range, "Top", 0) or 0)
    width = max(float(getattr(used_range, "Width", 0) or 0), 680)
    height = max(float(getattr(used_range, "Height", 0) or 0), 460)
    text_width = max(240, min(520, len(options.watermark_text) * options.watermark_font_size * 1.8))
    text_height = max(60, options.watermark_font_size * 2.4)
    horizontal_factor, vertical_step = SPACING_FACTORS.get(options.watermark_spacing, SPACING_FACTORS["medium"])
    x_step = text_width * horizontal_factor
    y_step = max(vertical_step, text_height * 1.8)
    rgb = excel_rgb(options.watermark_color)
    transparency = max(0.0, min(1.0, 1 - options.watermark_opacity / 100))

    shape_count = 0
    y = top
    row = 0
    while y <= top + height and shape_count < 90:
        x = left - (text_width * 0.4 if row % 2 else 0)
        while x <= left + width and shape_count < 90:
            shape = sheet.Shapes.AddTextbox(MSO_TEXT_ORIENTATION_HORIZONTAL, x, y, text_width, text_height)
            configure_watermark_shape(shape, options, rgb, transparency)
            shape_count += 1
            x += x_step
        y += y_step
        row += 1


def configure_watermark_shape(shape, options: ExcelPreviewOptions, rgb: int, transparency: float):
    shape.Name = f"PreviewWatermark_{uuid4().hex[:8]}"
    shape.Rotation = options.watermark_rotation
    shape.Locked = True
    try:
        shape.Placement = 3
    except Exception:
        pass
    try:
        shape.Fill.Visible = MSO_FALSE
        shape.Line.Visible = MSO_FALSE
    except Exception:
        pass
    try:
        text_range = shape.TextFrame2.TextRange
        text_range.Text = options.watermark_text
        text_range.Font.Size = options.watermark_font_size
        text_range.Font.Bold = MSO_TRUE
        text_range.Font.Fill.ForeColor.RGB = rgb
        text_range.Font.Fill.Transparency = transparency
        shape.TextFrame2.VerticalAnchor = 3
        shape.TextFrame2.TextRange.ParagraphFormat.Alignment = 2
    except Exception:
        shape.TextFrame.Characters().Text = options.watermark_text
        shape.TextFrame.Characters().Font.Size = options.watermark_font_size
        shape.TextFrame.Characters().Font.Color = rgb


def process_with_openpyxl(output_excel: Path, options: ExcelPreviewOptions, progress_callback=None) -> dict:
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image as OpenpyxlImage
    except Exception as exc:
        raise ProcessingError(f"缺少 openpyxl 依赖：{exc}") from exc

    if output_excel.suffix.lower() == ".xlsm":
        keep_vba = True
    else:
        keep_vba = False

    formula_wb = load_workbook(output_excel, keep_vba=keep_vba)
    data_wb = load_workbook(output_excel, data_only=True, keep_vba=keep_vba)
    processed_sheets = []
    skipped_hidden_sheets = []
    formulas_detected_count = 0
    formulas_converted_count = 0
    missing_cached_formulas = []
    watermark_sheets_count = 0
    protected_sheets_count = 0
    watermark_image = None
    warnings = ["当前环境可能无法准确将公式转为数值，建议安装 Excel/WPS 或先手动打开保存一次。"]

    if options.add_watermark:
        watermark_image = create_watermark_png(output_excel.parent, options)

    total_sheets = len(formula_wb.worksheets)
    for sheet_index, sheet in enumerate(formula_wb.worksheets, start=1):
        report_progress(progress_callback, f"正在处理工作表：{sheet.title}", sheet_index, total_sheets)
        if sheet.sheet_state != "visible" and not options.include_hidden_sheets:
            skipped_hidden_sheets.append(sheet.title)
            continue
        processed_sheets.append(sheet.title)
        data_sheet = data_wb[sheet.title]
        if options.convert_formulas:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f" or (isinstance(cell.value, str) and cell.value.startswith("=")):
                        formulas_detected_count += 1
                        cached = data_sheet[cell.coordinate].value
                        if cached is None:
                            missing_cached_formulas.append(f"{sheet.title}!{cell.coordinate}")
                            continue
                        cell.value = cached
                        formulas_converted_count += 1

        if options.add_watermark and watermark_image:
            sheet.add_image(OpenpyxlImage(str(watermark_image)), "A1")
            watermark_sheets_count += 1

        if options.protect_sheets:
            for row in sheet.iter_rows():
                for cell in row:
                    cell.protection = cell.protection.copy(locked=True)
            protect_openpyxl_sheet(sheet, options.protection_password)
            protected_sheets_count += 1

    if missing_cached_formulas:
        raise ProcessingError("当前环境无法准确计算公式，请先用 WPS/Excel 打开原文件并保存一次，或安装 Excel/WPS 自动化支持。")

    workbook_structure_protected = False
    if options.protect_workbook_structure:
        formula_wb.security.lockStructure = True
        formula_wb.security.workbookPassword = options.protection_password
        workbook_structure_protected = True

    formula_wb.save(output_excel)
    return {
        "total_sheets": len(formula_wb.worksheets),
        "processed_sheets": processed_sheets,
        "skipped_hidden_sheets": skipped_hidden_sheets,
        "processed_sheets_count": len(processed_sheets),
        "skipped_hidden_sheets_count": len(skipped_hidden_sheets),
        "preview_security_mode": "locked_excel",
        "preview_security_mode_label": "普通锁定版 Excel",
        "image_based_sheets_count": 0,
        "formulas_detected_count": formulas_detected_count,
        "formulas_converted_count": formulas_converted_count,
        "formula_conversion_required": options.convert_formulas,
        "formula_conversion_status": formula_conversion_status(options.convert_formulas, formulas_detected_count, formulas_converted_count, True),
        "formula_conversion_engine": "openpyxl" if options.convert_formulas else "disabled",
        "formula_conversion_note": "未检测到公式单元格。" if formulas_detected_count == 0 else "",
        "watermark_enabled": options.add_watermark,
        "watermark_text": options.watermark_text,
        "watermark_style": watermark_style_dict(options),
        "watermark_sheets_count": watermark_sheets_count,
        "protected_sheets_count": protected_sheets_count,
        "workbook_structure_protected": workbook_structure_protected,
        "edit_protection_enabled": options.protect_sheets or options.protect_workbook_structure,
        "real_cell_data_removed": False,
        "formula_removed": options.convert_formulas and formulas_converted_count == formulas_detected_count,
        "copy_risk_level": "medium",
        "screenshot_engine": "not_used",
        "images_preserved_in_screenshot": "unknown",
        "charts_preserved_in_screenshot": "unknown",
        "risk_notice": "该文件已保护工作表，但 Excel 工作表保护不是强加密，仍可能被复制或绕过。如需减少复制风险，请使用图片化防复制版。",
        "processing_engine": "openpyxl",
        "warnings": warnings + ["当前环境可能无法完全禁止选择和复制锁定单元格。"],
        "errors": [],
    }


def count_source_formulas(workbook_path: Path, sheet_names: set[str]) -> int:
    try:
        from openpyxl import load_workbook
    except Exception:
        return 0

    try:
        workbook = load_workbook(workbook_path, data_only=False, read_only=True, keep_vba=workbook_path.suffix.lower() == ".xlsm")
    except Exception:
        return 0

    try:
        count = 0
        for sheet in workbook.worksheets:
            if sheet.title not in sheet_names:
                continue
            for row in sheet.iter_rows():
                for cell in row:
                    value = cell.value
                    if cell.data_type == "f" or (isinstance(value, str) and value.startswith("=")):
                        count += 1
        return count
    finally:
        workbook.close()


def protect_openpyxl_sheet(sheet, password: str):
    sheet.protection.sheet = True
    sheet.protection.password = password
    sheet.protection.objects = True
    sheet.protection.scenarios = True
    sheet.protection.formatCells = True
    sheet.protection.formatColumns = True
    sheet.protection.formatRows = True
    sheet.protection.insertColumns = True
    sheet.protection.insertRows = True
    sheet.protection.deleteColumns = True
    sheet.protection.deleteRows = True
    sheet.protection.selectLockedCells = False
    sheet.protection.selectUnlockedCells = False


def dedupe_texts(items: list[str]) -> list[str]:
    deduped = []
    for item in items:
        if item and item not in deduped:
            deduped.append(item)
    return deduped


def screenshot_engine_from_engines(engines: set[str]) -> str:
    if "Excel COM" in engines:
        return "Excel COM"
    if "WPS COM" in engines:
        return "WPS COM"
    if engines:
        return "fallback_openpyxl"
    return "unknown"


def formula_conversion_status(enabled: bool, detected: int, converted: int, success: bool) -> str:
    if not enabled:
        return "disabled"
    if not success:
        return "failed"
    if converted == detected:
        return "converted"
    return "failed"


def report_progress(progress_callback, message: str, current: int | None = None, total: int | None = None):
    if not progress_callback:
        return
    try:
        progress_callback(message, current, total)
    except TypeError:
        progress_callback(message)


def create_watermark_png(folder: Path, options: ExcelPreviewOptions) -> Path:
    path = folder / f"watermark_{uuid4().hex[:8]}.png"
    image = Image.new("RGBA", (1000, 700), (255, 255, 255, 0))
    layer = Image.new("RGBA", image.size, (255, 255, 255, 0))
    draw = ImageDraw.Draw(layer)
    font = ImageFont.load_default()
    rgba = hex_to_rgba(options.watermark_color, options.watermark_opacity)
    for y in range(0, image.height, 160):
        for x in range(-120, image.width, 360):
            draw.text((x, y), options.watermark_text, fill=rgba, font=font)
    rotated = layer.rotate(options.watermark_rotation, expand=False)
    image.alpha_composite(rotated)
    image.save(path)
    return path


def list_workbook_sheets(path: Path) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ProcessingError(f"缺少 openpyxl 依赖，无法读取工作表列表：{exc}") from exc

    try:
        wb = load_workbook(path, read_only=False, keep_vba=path.suffix.lower() == ".xlsm")
    except Exception as exc:
        raise ProcessingError(f"读取 Excel 工作表失败：{exc}") from exc
    try:
        sheets = []
        for index, sheet in enumerate(wb.worksheets, start=1):
            state = sheet.sheet_state or "visible"
            print_area = ""
            try:
                print_area = str(sheet.print_area or "")
            except Exception:
                print_area = ""
            sheets.append(
                {
                    "index": index,
                    "name": sheet.title,
                    "visible": state == "visible",
                    "state": state,
                    "label": sheet.title if state == "visible" else f"{sheet.title}（隐藏）",
                    "print_area": print_area,
                }
            )
        return sheets
    finally:
        wb.close()


def read_upload_meta(upload_id: str, output_root: Path) -> dict:
    upload_dir = resolve_upload_dir(upload_id, output_root)
    meta_path = upload_dir / "upload_meta.json"
    if not meta_path.exists():
        raise ProcessingError("当前文件不存在，请重新上传 Excel。")
    with meta_path.open("r", encoding="utf-8") as file:
        return json.load(file)


def upload_source_path(upload_meta: dict, output_root: Path) -> Path:
    upload_dir = resolve_upload_dir(upload_meta["upload_id"], output_root)
    source_path = upload_dir / upload_meta["source_file"]
    if not source_path.exists():
        raise ProcessingError("当前文件不存在，请重新上传 Excel。")
    return source_path


def resolve_upload_dir(upload_id: str, output_root: Path) -> Path:
    if "/" in upload_id or "\\" in upload_id or not upload_id.startswith("excel_preview_"):
        raise ProcessingError("Excel 临时任务 ID 无效。")
    root = (output_root / TEMP_PREVIEW_DIRNAME).resolve()
    upload_dir = (root / upload_id).resolve()
    if root not in upload_dir.parents:
        raise ProcessingError("Excel 临时任务路径无效。")
    return upload_dir


def generate_preview_with_com(
    workbook_path: Path,
    sheet_name: str,
    range_type: str,
    options: ExcelPreviewOptions,
    preview_path: Path,
    prefer_wps: bool,
    prefer_silent: bool = False,
) -> dict:
    try:
        import pythoncom
        import win32com.client
    except Exception as exc:
        raise ProcessingError(f"Excel/WPS COM 不可用：{exc}") from exc

    workbook_path = Path(workbook_path).resolve()
    preview_path = Path(preview_path).resolve()
    preview_path.parent.mkdir(parents=True, exist_ok=True)
    engine_name = "WPS COM" if prefer_wps else "Excel COM"
    app = None
    workbook = None
    app_process_ids: set[int] = set()
    office_processes_before = office_process_snapshot()
    pythoncom.CoInitialize()
    try:
        app = win32com.client.DispatchEx("Ket.Application" if prefer_wps else "Excel.Application")
        app_process_ids = com_app_process_ids(app)
        configure_com_capture_window(app, "quiet" if prefer_silent else "visible")
        workbook = open_com_workbook(app, workbook_path, read_only=True)
        configure_active_com_window(app, "quiet" if prefer_silent else "visible")
        return capture_open_workbook_sheet_to_png(engine_name, app, workbook, sheet_name, range_type, options, preview_path)
    except ProcessingError:
        raise
    except Exception as exc:
        message = str(exc)
        if looks_like_locked_file_error(message):
            raise ProcessingError("文件正在被 WPS/Excel 占用，请先关闭后再预览。") from exc
        raise ProcessingError(f"{engine_name} 截图预览失败：{message}") from exc
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=False)
            except Exception:
                pass
        if app is not None:
            cleanup_com_application(app, app_process_ids)
        cleanup_new_office_processes(office_processes_before)
        pythoncom.CoUninitialize()


def capture_open_workbook_sheet_to_png(engine_name: str, app, workbook, sheet_name: str, range_type: str, options: ExcelPreviewOptions, output_path: Path, window_mode: str | None = None) -> dict:
    output_path = Path(output_path).resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_base = output_path.with_name(f"range_base_{uuid4().hex[:8]}.png")
    temp_pdf = output_path.with_name(f"range_base_{uuid4().hex[:8]}.pdf")
    try:
        sheet = get_com_worksheet(workbook, sheet_name)
        sheet.Activate()
        if window_mode:
            configure_active_com_window(app, window_mode)
        try:
            app.ActiveWindow.Activate()
        except Exception:
            pass
        target_range = preview_target_range(sheet, range_type)
        width = max(240.0, float(target_range.Width))
        height = max(160.0, float(target_range.Height))
        capture_method = "copy_picture_clipboard"
        try:
            capture_range_with_clipboard(app, sheet, target_range, temp_base)
        except ProcessingError as clipboard_exc:
            try:
                export_range_to_pdf_image(sheet, target_range, range_type, temp_pdf, temp_base)
                capture_method = "pdf_export"
            except ProcessingError as pdf_exc:
                raise ProcessingError(f"{engine_name} 真实截图失败：{clipboard_exc}；PDF 渲染也失败：{pdf_exc}") from pdf_exc
        if preview_image_looks_blank(temp_base):
            raise ProcessingError(f"{engine_name} 截图预览未截取到工作表内容。")
        overlay_watermark(temp_base, output_path, options)
        ensure_png_created(output_path)
        return {
            "preview_engine": engine_name,
            "screenshot_engine": engine_name,
            "capture_method": capture_method,
            "range_type": range_type,
            "pixel_size": {"width": int(width), "height": int(height)},
            "warnings": [],
        }
    finally:
        if temp_base.exists():
            try:
                temp_base.unlink()
            except OSError:
                pass
        if temp_pdf.exists():
            try:
                temp_pdf.unlink()
            except OSError:
                pass


def open_com_workbook(app, workbook_path: Path, read_only: bool):
    workbook_path = Path(workbook_path).resolve()
    filename = str(workbook_path)
    attempts = [
        lambda: app.Workbooks.Open(
            filename,
            UpdateLinks=0,
            ReadOnly=read_only,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
        ),
        lambda: app.Workbooks.Open(filename, 0, read_only),
        lambda: app.Workbooks.Open(Filename=filename, ReadOnly=read_only),
        lambda: app.Workbooks.Open(filename),
    ]
    last_error = None
    for attempt in attempts:
        try:
            return attempt()
        except Exception as exc:
            last_error = exc
    raise last_error


def get_com_worksheet(workbook, sheet_name: str):
    target = str(sheet_name).strip()
    count = None
    last_error = None
    for _ in range(15):
        try:
            count = int(workbook.Worksheets.Count)
            break
        except Exception as exc:
            last_error = exc
            time.sleep(0.2)
    if count is None:
        raise ProcessingError("无法读取工作表列表。") from last_error
    for index in range(1, count + 1):
        try:
            sheet = workbook.Worksheets(index)
            if str(sheet.Name).strip() == target:
                return sheet
        except Exception:
            continue
    raise ProcessingError("当前工作表不存在。")


def capture_range_with_clipboard(app, sheet, target_range, output_path: Path):
    try:
        try:
            import win32clipboard

            win32clipboard.OpenClipboard()
            win32clipboard.EmptyClipboard()
            win32clipboard.CloseClipboard()
        except Exception:
            pass
        sheet.Activate()
        try:
            app.ActiveWindow.Zoom = 100
        except Exception:
            pass
        target_range.Select()
        time.sleep(0.25)
        target_range.CopyPicture(Appearance=XL_SCREEN, Format=XL_BITMAP)
        image = None
        for _ in range(12):
            time.sleep(0.15)
            image = ImageGrab.grabclipboard()
            if isinstance(image, Image.Image):
                break
        if not isinstance(image, Image.Image):
            raise ProcessingError("剪贴板没有返回图片。")
        image.convert("RGB").save(output_path, "PNG")
    except ProcessingError:
        raise
    except Exception as exc:
        raise ProcessingError(f"CopyPicture 截图失败：{exc}") from exc


def export_range_to_pdf_image(sheet, target_range, range_type: str, pdf_path: Path, output_path: Path):
    original_print_area = ""
    original_zoom = None
    original_fit_wide = None
    original_fit_tall = None
    try:
        page_setup = sheet.PageSetup
        original_print_area = str(page_setup.PrintArea or "")
        try:
            original_zoom = page_setup.Zoom
            original_fit_wide = page_setup.FitToPagesWide
            original_fit_tall = page_setup.FitToPagesTall
        except Exception:
            pass
        if range_type != "print_area" or not original_print_area:
            page_setup.PrintArea = target_range.Address
        try:
            page_setup.Zoom = 100
            page_setup.FitToPagesWide = False
            page_setup.FitToPagesTall = False
        except Exception:
            pass
        sheet.ExportAsFixedFormat(Type=XL_TYPE_PDF, Filename=str(pdf_path), OpenAfterPublish=False)
        if not pdf_path.exists() or pdf_path.stat().st_size == 0:
            raise ProcessingError("临时 PDF 未生成。")
        render_pdf_to_png(pdf_path, output_path)
    except ProcessingError:
        raise
    except Exception as exc:
        raise ProcessingError(f"PDF 渲染截图失败：{exc}") from exc
    finally:
        try:
            sheet.PageSetup.PrintArea = original_print_area
            if original_zoom is not None:
                sheet.PageSetup.Zoom = original_zoom
            if original_fit_wide is not None:
                sheet.PageSetup.FitToPagesWide = original_fit_wide
            if original_fit_tall is not None:
                sheet.PageSetup.FitToPagesTall = original_fit_tall
        except Exception:
            pass


def render_pdf_to_png(pdf_path: Path, output_path: Path):
    try:
        import fitz
    except Exception as exc:
        raise ProcessingError(f"缺少 PyMuPDF，无法渲染临时 PDF：{exc}") from exc

    doc = fitz.open(pdf_path)
    try:
        page_images = []
        for page in doc:
            pixmap = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
            mode = "RGB" if pixmap.n < 4 else "RGBA"
            image = Image.frombytes(mode, [pixmap.width, pixmap.height], pixmap.samples).convert("RGB")
            page_images.append(image)
        if not page_images:
            raise ProcessingError("临时 PDF 没有页面。")
        if len(page_images) == 1:
            page_images[0].save(output_path, "PNG")
            return
        width = max(image.width for image in page_images)
        height = sum(image.height for image in page_images)
        stitched = Image.new("RGB", (width, height), "white")
        y = 0
        for image in page_images:
            stitched.paste(image, (0, y))
            y += image.height
        stitched.save(output_path, "PNG")
    finally:
        doc.close()


def preview_target_range(sheet, range_type: str):
    if range_type in {"print_area", "auto"}:
        try:
            print_area = str(sheet.PageSetup.PrintArea or "")
            if print_area:
                return expand_range_to_visible_objects(sheet, sheet.Range(print_area))
        except Exception:
            pass
    return expand_range_to_visible_objects(sheet, sheet.UsedRange)


def expand_range_to_visible_objects(sheet, base_range):
    try:
        min_row = int(base_range.Row)
        min_col = int(base_range.Column)
        max_row = min_row + int(base_range.Rows.Count) - 1
        max_col = min_col + int(base_range.Columns.Count) - 1
    except Exception:
        return base_range

    try:
        shapes = sheet.Shapes
        shape_count = int(shapes.Count)
    except Exception:
        shape_count = 0

    for index in range(1, shape_count + 1):
        try:
            shape = shapes.Item(index)
            if hasattr(shape, "Visible") and int(shape.Visible) == MSO_FALSE:
                continue
            top_left = shape.TopLeftCell
            bottom_right = shape.BottomRightCell
            min_row = min(min_row, int(top_left.Row))
            min_col = min(min_col, int(top_left.Column))
            max_row = max(max_row, int(bottom_right.Row))
            max_col = max(max_col, int(bottom_right.Column))
        except Exception:
            continue

    try:
        return sheet.Range(sheet.Cells(min_row, min_col), sheet.Cells(max_row, max_col))
    except Exception:
        return base_range


def generate_preview_with_openpyxl(
    workbook_path: Path,
    sheet_name: str,
    range_type: str,
    options: ExcelPreviewOptions,
    preview_path: Path,
) -> dict:
    workbook_path = Path(workbook_path).resolve()
    preview_path = Path(preview_path).resolve()
    preview_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise ProcessingError(f"openpyxl/Pillow 近似预览不可用：{exc}") from exc

    wb = load_workbook(workbook_path, data_only=True, read_only=False, keep_vba=workbook_path.suffix.lower() == ".xlsm")
    try:
        if sheet_name not in wb.sheetnames:
            raise ProcessingError("当前工作表不存在。")
        sheet = wb[sheet_name]
        max_row = min(sheet.max_row or 1, 80)
        max_col = min(sheet.max_column or 1, 24)
        cell_width = 96
        cell_height = 28
        image = Image.new("RGB", (max_col * cell_width + 1, max_row * cell_height + 1), "white")
        draw = ImageDraw.Draw(image)
        font = load_watermark_font(12)
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                x0 = (col - 1) * cell_width
                y0 = (row - 1) * cell_height
                x1 = x0 + cell_width
                y1 = y0 + cell_height
                draw.rectangle((x0, y0, x1, y1), outline=(220, 226, 235))
                value = sheet.cell(row=row, column=col).value
                if value is not None:
                    draw.text((x0 + 4, y0 + 6), str(value)[:16], fill=(31, 41, 51), font=font)
        base_path = preview_path.with_name(f"openpyxl_base_{uuid4().hex[:8]}.png")
        image.save(base_path)
        overlay_watermark(base_path, preview_path, options)
        base_path.unlink(missing_ok=True)
        return {
            "preview_engine": "openpyxl + Pillow",
            "screenshot_engine": "fallback_openpyxl",
            "capture_method": "fallback_openpyxl",
            "range_type": range_type,
            "warnings": [APPROXIMATE_FALLBACK_WARNING],
        }
    finally:
        wb.close()


def overlay_watermark(base_path: Path, output_path: Path, options: ExcelPreviewOptions):
    if not options.add_watermark:
        shutil.copy2(base_path, output_path)
        return
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        with Image.open(base_path) as source:
            image = source.convert("RGBA")
    layer = Image.new("RGBA", image.size, (255, 255, 255, 0))
    tile = watermark_text_tile(options)
    horizontal_factor, vertical_step = SPACING_FACTORS.get(options.watermark_spacing, SPACING_FACTORS["medium"])
    x_step = max(120, int(tile.width * horizontal_factor))
    y_step = max(80, int(max(vertical_step, tile.height * 0.85)))
    y = -tile.height
    row = 0
    while y < image.height + tile.height:
        x = -tile.width - (tile.width // 2 if row % 2 else 0)
        while x < image.width + tile.width:
            layer.alpha_composite(tile, (int(x), int(y)))
            x += x_step
        y += y_step
        row += 1
    composed = Image.alpha_composite(image, layer).convert("RGB")
    composed.save(output_path, "PNG")


def preview_image_looks_blank(path: Path) -> bool:
    try:
        with Image.open(path) as source:
            image = source.convert("RGB")
            if image.width > 80 and image.height > 80:
                left = int(image.width * 0.02)
                top = int(image.height * 0.08)
                right = int(image.width * 0.98)
                bottom = int(image.height * 0.92)
                image = image.crop((left, top, right, bottom))
            image.thumbnail((480, 480))
            pixels = list(image.getdata())
    except Exception:
        return False
    if not pixels:
        return True
    non_white = 0
    for red, green, blue in pixels[:: max(1, len(pixels) // 12000)]:
        if red < 245 or green < 245 or blue < 245:
            non_white += 1
    checked = max(1, len(pixels[:: max(1, len(pixels) // 12000)]))
    return non_white / checked < 0.003


def watermark_text_tile(options: ExcelPreviewOptions) -> Image.Image:
    font = load_watermark_font(options.watermark_font_size)
    text = options.watermark_text or "保密文件"
    text_image = Image.new("RGBA", (10, 10), (255, 255, 255, 0))
    draw = ImageDraw.Draw(text_image)
    bbox = draw.textbbox((0, 0), text, font=font)
    width = max(240, bbox[2] - bbox[0] + 80)
    height = max(80, bbox[3] - bbox[1] + 50)
    text_image = Image.new("RGBA", (width, height), (255, 255, 255, 0))
    draw = ImageDraw.Draw(text_image)
    rgba = hex_to_rgba(options.watermark_color, options.watermark_opacity)
    draw.text((40, height // 2 - (bbox[3] - bbox[1]) // 2), text, fill=rgba, font=font)
    return text_image.rotate(options.watermark_rotation, expand=True, resample=Image.Resampling.BICUBIC)


def load_watermark_font(size: int):
    candidates = [
        Path(r"C:\Windows\Fonts\msyh.ttc"),
        Path(r"C:\Windows\Fonts\simhei.ttf"),
        Path(r"C:\Windows\Fonts\arial.ttf"),
    ]
    for path in candidates:
        if path.exists():
            try:
                return ImageFont.truetype(str(path), size=size)
            except Exception:
                continue
    return ImageFont.load_default()


def prefixed_preview_stem(safe_stem: str) -> str:
    return safe_stem if safe_stem.startswith(PREVIEW_PREFIX) else f"{PREVIEW_PREFIX}{safe_stem}"


def watermark_style_dict(options: ExcelPreviewOptions) -> dict:
    return {
        "font_size": options.watermark_font_size,
        "color": options.watermark_color,
        "opacity_percent": options.watermark_opacity,
        "rotation": options.watermark_rotation,
        "spacing": options.watermark_spacing,
    }


def excel_rgb(hex_color: str) -> int:
    value = (hex_color or "#B8B8B8").lstrip("#")
    if len(value) != 6:
        value = "B8B8B8"
    red = int(value[0:2], 16)
    green = int(value[2:4], 16)
    blue = int(value[4:6], 16)
    return red + (green << 8) + (blue << 16)


def hex_to_rgba(hex_color: str, opacity: int) -> tuple[int, int, int, int]:
    value = (hex_color or "#B8B8B8").lstrip("#")
    if len(value) != 6:
        value = "B8B8B8"
    red = int(value[0:2], 16)
    green = int(value[2:4], 16)
    blue = int(value[4:6], 16)
    alpha = int(max(0, min(100, opacity)) / 100 * 255)
    return red, green, blue, alpha


def looks_like_locked_file_error(message: str) -> bool:
    lowered = message.lower()
    return any(token in lowered for token in ["permission", "access", "locked", "read-only", "只读", "权限", "正在使用"])


def should_try_wps_preview(message: str) -> bool:
    lowered = message.lower()
    return any(
        token in lowered
        for token in [
            "class not registered",
            "invalid class string",
            "无效类字符串",
            "未注册",
            "不可用",
            "workbooks 的 open",
            "open 方法无效",
            "方法无效",
        ]
    )


def options_from_form(form) -> ExcelPreviewOptions:
    security_mode = (form.get("preview_security_mode") or "image_based").strip()
    if security_mode not in {"image_based", "locked_excel"}:
        security_mode = "image_based"
    image_range_type = (form.get("image_range_type") or form.get("range_type") or "auto").strip()
    if image_range_type not in {"auto", "used_range", "print_area"}:
        image_range_type = "auto"
    output_mode = (form.get("output_mode") or "desktop").strip()
    if output_mode not in {"desktop", "app_output"}:
        output_mode = "desktop"
    screenshot_window_mode = (form.get("screenshot_window_mode") or "quiet").strip()
    if screenshot_window_mode not in {"quiet", "visible"}:
        screenshot_window_mode = "quiet"
    convert_formulas = parse_bool(form.get("convert_formulas"), True)
    if security_mode == "image_based":
        convert_formulas = False
    return ExcelPreviewOptions(
        watermark_text=(form.get("watermark_text") or "保密文件").strip() or "保密文件",
        watermark_font_size=clamp_int(form.get("watermark_font_size"), 12, 80, 28),
        watermark_color=(form.get("watermark_color") or "#B8B8B8").strip() or "#B8B8B8",
        watermark_opacity=clamp_int(form.get("watermark_opacity"), 5, 80, 20),
        watermark_rotation=clamp_int(form.get("watermark_rotation"), -90, 90, -30),
        watermark_spacing=(form.get("watermark_spacing") or "medium").strip() or "medium",
        protection_password=form.get("protection_password") or "123456",
        convert_formulas=convert_formulas,
        add_watermark=parse_bool(form.get("add_watermark"), True),
        protect_sheets=parse_bool(form.get("protect_sheets"), True),
        protect_workbook_structure=parse_bool(form.get("protect_workbook_structure"), True),
        include_hidden_sheets=parse_bool(form.get("include_hidden_sheets"), False),
        preview_security_mode=security_mode,
        image_range_type=image_range_type,
        allow_approximate_fallback=parse_bool(form.get("allow_approximate_fallback"), False),
        output_mode=output_mode,
        output_report_to_desktop=parse_bool(form.get("output_report_to_desktop"), False),
        screenshot_window_mode=screenshot_window_mode,
    )


def parse_bool(value, default: bool) -> bool:
    if value is None:
        return default
    return str(value).lower() in {"1", "true", "yes", "on", "是"}


def clamp_int(value, minimum: int, maximum: int, default: int) -> int:
    try:
        number = int(value)
    except (TypeError, ValueError):
        return default
    return max(minimum, min(maximum, number))
